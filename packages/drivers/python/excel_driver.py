"""excel driver — bulk xlsx approach via xlwings.

strategy:
  1. one xlsx with a sheet per task (grid at A1-area, formula at AA1)
  2. open in excel ONCE via xlwings, full recalc, save (caches values), quit
  3. read cached values back via openpyxl `data_only=True` — no further round-trip

dynamic-array formulas spill to their output shape when excel 2021/365 saves
the workbook; older excel truncates spills to the top-left.

usage:
  python excel_driver.py <task.json> <result.json>
  python excel_driver.py --batch <tasks.json> <results.json>
  python excel_driver.py --quit
  python excel_driver.py [--workbook <path>] ...   # accepted but ignored
"""

from __future__ import annotations

import json
import os
import re
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Any
from xml.etree import ElementTree as ET

import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText


# Excel 1900 date system (matches Sheets), day 0 = 1899-12-30.
# Excel-for-Mac historically defaulted to 1904 date system; openpyxl exposes
# the workbook's actual epoch via wb.epoch (datetime(1899,12,30) for 1900-mode,
# datetime(1904,1,1) for 1904-mode). _dt_to_serial accepts the epoch as a
# parameter so 1904-mode workbooks convert correctly. (Was: hardcoded 1900 →
# silently wrong serials by 1462 days when the workbook was 1904-mode.)
_DATE_EPOCH = datetime(1899, 12, 30)


def _dt_to_serial(v: Any, epoch: datetime = _DATE_EPOCH) -> float:
    # datetime / date / time → excel serial; matches sheets' UNFORMATTED output.
    # `epoch` defaults to 1900-mode but should be passed from wb.epoch when
    # the workbook is in 1904 mode.
    if isinstance(v, datetime):
        return (v - epoch).total_seconds() / 86400.0
    if isinstance(v, date):
        return (datetime(v.year, v.month, v.day) - epoch).days
    if isinstance(v, time):
        return (v.hour * 3600 + v.minute * 60 + v.second + v.microsecond / 1e6) / 86400.0
    raise TypeError(type(v))


VERBOSE = os.environ.get("ASSAY_VERBOSE", "") == "1"

# AA1, 1-indexed — cols A-Z reserved for grid: cells. These are the LEGACY
# default placement (formula cell + read window); the per-task `placement` from
# the TS packing plan overrides them (see _placement).
TARGET_ROW = 1
TARGET_COL = 27
SPILL_ROWS = 20
SPILL_COLS = 20


def _placement(task: dict[str, Any]) -> tuple[int, int, int, int]:
    """(top, left, rows, cols), 1-indexed — the task's formula cell + read window.

    Comes from the TS-computed packing plan (`placement`); the geometry is
    single-sourced there (packing.ts) so this is a dumb executor. Falls back to the
    legacy AA1 / 20×20 window when absent (skip tasks, or an unplanned caller).
    """
    p = task.get("placement")
    if isinstance(p, dict):
        return (int(p["top"]), int(p["left"]), int(p["rows"]), int(p["cols"]))
    return (TARGET_ROW, TARGET_COL, SPILL_ROWS, SPILL_COLS)

_CELL_REF_RE = re.compile(r"^([A-Z]+)(\d+)$")
_ERROR_STR_RE = re.compile(r"^#[A-Z0-9/]+!?\??$")


def log(msg: str) -> None:
    if VERBOSE:
        print(f"[excel] {msg}", file=sys.stderr, flush=True)


def parse_cell_ref(ref: str) -> tuple[int, int]:
    m = _CELL_REF_RE.match(ref.upper())
    if not m:
        raise ValueError(f"Invalid cell reference: {ref}")
    letters, row = m.group(1), int(m.group(2))
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


def col_letters(col: int) -> str:
    out = ""
    while col > 0:
        col, r = divmod(col - 1, 26)
        out = chr(65 + r) + out
    return out


def _host_key(task: dict[str, Any], index: int) -> tuple[str, int]:
    """The co-hosting key: ("h", host) for a planned (possibly co-tiled) task, or a
    unique ("u", index) for an unplanned one (so it lands on its own sheet)."""
    p = task.get("placement")
    if isinstance(p, dict) and "host" in p:
        return ("h", int(p["host"]))
    return ("u", index)


def build_workbook(tasks: list[dict[str, Any]], path: str) -> list[str]:
    # One sheet per distinct HOST (co-hosted lump tasks share a sheet; in-place/
    # isolate tasks are solo per the plan). Only grids are written here — formulas
    # are entered later via xlwings `Range.formula2` (the trigger for dynamic-array
    # spill; openpyxl-written formulas come out legacy non-array, top-left only).
    # Only solo (in-place/isolate) tasks carry grids, so no seed collides with a
    # co-tenant. Returns sheet_of_task[i] = the sheet task i lives on (dups where
    # co-hosted), so recalc/read place + read each task at its own tile.
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    sheet_of_task: list[str] = []
    key_to_sheet: dict[tuple[str, int], str] = {}
    for i, task in enumerate(tasks):
        key = _host_key(task, i)
        name = key_to_sheet.get(key)
        if name is None:
            name = f"s{len(key_to_sheet)}"
            key_to_sheet[key] = name
            wb.create_sheet(title=name)
        sheet_of_task.append(name)

    for i, task in enumerate(tasks):
        ws = wb[sheet_of_task[i]]
        for ref, val in (task.get("grid") or {}).items():
            row, col = parse_cell_ref(ref)
            if val is None:
                continue
            cell = ws.cell(row=row, column=col)
            if isinstance(val, dict) and "error" in val:
                # D6: seed a REAL cell error literal — Excel reads ISERROR(A1)=TRUE.
                cell.value = val["error"]
                cell.data_type = "e"
            elif isinstance(val, str):
                # D1: a string is TEXT, even when it LOOKS like an error ("#DIV/0!")
                # or a number ("3"). openpyxl auto-infers 'e' for error-code strings,
                # so set the value then force 's' to keep it text (ISTEXT(A1)=TRUE).
                cell.value = val
                cell.data_type = "s"
            else:
                # number / bool — openpyxl infers 'n' / 'b' faithfully.
                cell.value = val

    wb.save(path)
    return sheet_of_task


def _configure_app(app: Any) -> None:
    for attr, value in [("display_alerts", False), ("screen_updating", False)]:
        try:
            setattr(app, attr, value)
        except Exception:
            pass


_IS_DARWIN = sys.platform == "darwin"


def _api_attr(obj: Any, *names: str) -> Any:
    """Look up an `.api` attribute under the first name that resolves.

    Windows xlwings uses pywin32 COM (CamelCase like `Value2`); Mac xlwings
    uses appscript with lowercase AppleScript names (`value2`). On Mac the
    return is a lazy `Reference` — caller must explicitly call `.get()` to
    fetch the concrete value (and pay the Apple Event cost). Do NOT auto-fetch
    here: even one extra round-trip per cell crashes Excel at hundreds-of-cell
    scale.
    """
    last_exc: Exception | None = None
    for name in names:
        try:
            v = obj
            for part in name.split("."):
                v = getattr(v, part)
            return v
        except Exception as e:  # noqa: BLE001
            last_exc = e
            continue
    if last_exc is not None:
        raise last_exc
    return None


def _is_scalar_value(v: Any) -> bool:
    """True for JSON-friendly primitives + datetime. False for appscript
    Reference objects / COM error variants / other engine-internal types
    that don't round-trip via str()."""
    if v is None or isinstance(v, (bool, int, float, str)):
        return True
    if isinstance(v, (datetime, date, time)):
        return True
    return False


def _capture_surface_b_for_sheet(
    sht: Any,
    top: int = TARGET_ROW,
    left: int = TARGET_COL,
    rows: int = SPILL_ROWS,
    cols: int = SPILL_COLS,
) -> dict[tuple[int, int], SurfaceB]:
    """Read Surface B (Value2 / DisplayFormat / SavedAsArray) per cell in the
    rows × cols read window anchored at (top, left).

    Platform notes:
    - **Value2**: Windows COM `Value2` ⇄ Mac AS `value2`. Bulk-read for the
      whole region in one call when supported.
    - **DisplayFormat.NumberFormat**: Windows COM `DisplayFormat.NumberFormat`
      ⇄ Mac AS `display_format.number_format`.
    - **SavedAsArray**: Windows-COM-only. Mac AS has `has_array` ("is this cell
      part of an array formula") which is semantically different — we still
      capture it as a best-effort signal but consumers should treat the
      Windows and Mac fields as distinct meanings.

    Each property is guarded — bridge differences leave fields as None rather
    than crashing. Surface B is best-effort; a missing field is not an error.
    """
    # Mac AE bridge is incompatible with Surface B capture: empirical finding
    # 2026-05-23 is that even a SINGLE bulk Value2 `.get()` per sheet crashes
    # Excel after 2-3 invocations (OSERROR -609 "Connection is invalid",
    # followed by procNotFound on subsequent ops). The bridge can't sustain
    # the round-trip volume Surface B requires. Skip on Mac entirely; Windows
    # COM tolerates the call patterns. Future Mac Surface B work would need a
    # different bridge (Office.js, AppleScript bulk paths, etc.) — tracked in
    # the coalescing doc as a deferred follow-up.
    if _IS_DARWIN:
        return {}

    result: dict[tuple[int, int], SurfaceB] = {}
    # Bulk Value2 — one round trip for the whole region.
    bulk_value2: list[list[Any]] | None = None
    try:
        region = sht.range(
            (top, left),
            (top + rows - 1, left + cols - 1),
        )
        v2 = _api_attr(region.api, "Value2", "value2")
        if isinstance(v2, (list, tuple)):
            bulk_value2 = [list(row) if isinstance(row, (list, tuple)) else [row] for row in v2]
        else:
            bulk_value2 = [[v2]]
    except Exception as e:  # noqa: BLE001
        log(f"  Surface B Value2 bulk read failed: {e}")

    for r in range(top, top + rows):
        for c in range(left, left + cols):
            sb = SurfaceB()
            if bulk_value2 is not None:
                ri, ci = r - top, c - left
                if ri < len(bulk_value2) and ci < len(bulk_value2[ri]):
                    raw_v = bulk_value2[ri][ci]
                    if _is_scalar_value(raw_v):
                        sb.value2 = raw_v
            cell = sht.range((r, c))
            try:
                fmt = _api_attr(
                    cell.api,
                    "DisplayFormat.NumberFormat",
                    "display_format.number_format",
                )
                sb.display_number_format = str(fmt) if fmt is not None else None
            except Exception:  # noqa: BLE001
                pass
            try:
                sb.saved_as_array = bool(_api_attr(cell.api, "SavedAsArray", "has_array"))
            except Exception:  # noqa: BLE001
                pass
            if (
                sb.value2 is not None
                or sb.display_number_format is not None
                or sb.saved_as_array is not None
            ):
                result[(r, c)] = sb
    return result


def recalc_with_excel(
    app: Any,
    xlsx_path: str,
    placed: list[tuple[str, str, int, int, int, int]],
) -> dict[str, dict[tuple[int, int], SurfaceB]]:
    # open xlsx, enter each formula at its PLACEMENT cell via Range.formula2
    # (triggers dynamic-array spill for SORT/FILTER/TRANSPOSE/…), recalc, capture
    # Surface B (live `.api` properties — must happen BEFORE wb.save()/close
    # since the data only exists while the workbook is open), save, close.
    # placed[i] = (sheet_name, formula, top, left, rows, cols) — co-tiled tasks
    # share a sheet, each at its own (top,left) tile (build_workbook's grouping).
    # the app itself is left running so subsequent chunks don't re-trigger the
    # macos automation permission prompt.
    log(f"  opening {os.path.basename(xlsx_path)}...")
    wb = app.books.open(xlsx_path)
    # manual calc while entering formulas — otherwise every formula2 = ...
    # triggers a full-workbook recalc, scaling quadratically with sheet count
    # and blowing excel's memory budget past ~50 dynamic-array sheets.
    prior_calc = None
    try:
        prior_calc = app.calculation
        app.calculation = "manual"
    except Exception:
        pass
    try:
        log(f"  entering {len(placed)} formula(s) via Range.formula2...")
        for sheet_name, formula, top, left, _rows, _cols in placed:
            wb.sheets[sheet_name].range((top, left)).formula2 = formula
        log("  forcing calculation...")
        app.calculate()
    finally:
        if prior_calc is not None:
            try:
                app.calculation = prior_calc
            except Exception:
                pass
    log("  capturing Surface B (live `.api` properties)...")
    # One dict per sheet; co-tiled tasks' (non-overlapping) regions merge in.
    surface_b_by_sheet: dict[str, dict[tuple[int, int], SurfaceB]] = {}
    for sheet_name, _formula, top, left, rows, cols in placed:
        sht = wb.sheets[sheet_name]
        try:
            cap = _capture_surface_b_for_sheet(sht, top, left, rows, cols)
        except Exception as e:  # noqa: BLE001
            log(f"  Surface B capture failed for sheet {sheet_name}: {e}")
            cap = {}
        surface_b_by_sheet.setdefault(sheet_name, {}).update(cap)
    log("  saving workbook...")
    wb.save()
    wb.close()
    return surface_b_by_sheet


def cell_to_value(cell: Any, epoch: datetime = _DATE_EPOCH) -> Any:
    return _value_to_cell(cell.value, epoch)


def _value_to_cell(v: Any, epoch: datetime = _DATE_EPOCH) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if isinstance(v, float):
            if v != v:  # NaN
                return {"error": "#NUM!"}
            if v in (float("inf"), float("-inf")):
                return {"error": "#NUM!"}
            if v.is_integer() and abs(v) < 2**53:
                return int(v)
        return v
    if isinstance(v, str):
        if _ERROR_STR_RE.match(v):
            return {"error": v}
        return v
    if isinstance(v, (datetime, date, time)):
        # openpyxl returns python datetime objects for date-formatted cells;
        # normalise to excel/sheets serial so DATE/NOW/TODAY align with
        # gsheets' UNFORMATTED_VALUE readback. Uses the workbook's actual
        # epoch (1900-mode default; 1904-mode for Mac-legacy workbooks).
        serial = _dt_to_serial(v, epoch)
        if isinstance(serial, float) and serial.is_integer() and abs(serial) < 2**53:
            return int(serial)
        return serial
    return str(v)


@dataclass
class SurfaceB:
    """Live xlwings `.api.Range.*` properties captured during the Excel session.

    Per F21/F23/F24 of the audit: these are runtime properties Excel
    computes when the workbook is OPEN; openpyxl (data_only post-save)
    never sees them. Captured inside `recalc_with_excel` before
    `wb.save()` to avoid the workbook-closed trap.

    Notes for Mac vs Windows:
    - `.api.Value2`: bit-accurate raw value. Dates stay as numeric serials
      (no datetime conversion). On Windows it's the COM Value2 property;
      on Mac the Apple Events bridge exposes equivalent. Error variants
      (`CVErr`) come through as platform-specific types.
    - `.api.DisplayFormat.NumberFormat`: the post-conditional-formatting
      number format string. Closest Excel analog to gsheets effectiveFormat.
      May not exist on Mac Apple Events; failure is captured as None.
    - `.api.SavedAsArray`: the writer-side heuristic result for whether
      Excel will persist `<f t="array">` at save time. Boolean.

    All fields are best-effort: capture failures (Mac vs Windows API
    differences) leave None rather than crashing the driver.
    """

    value2: Any = None
    display_number_format: str | None = None
    saved_as_array: bool | None = None


@dataclass
class RawCellData:
    """Per-cell raw OOXML XML fields the openpyxl read drops or normalizes.

    Captures the structural axes the audit identified as load-bearing but
    invisible through openpyxl:

    - `cm`: cell metadata index — present on spill anchors (dynamic-array
      formulas) and Linked-Data-Type host cells. openpyxl drops this.
    - `vm`: value metadata index — Linked-Data-Type pointer into
      `xl/richData/`. openpyxl drops this.
    - `s`: style index (into styles.xml). openpyxl resolves it to
      number_format but loses the raw index.
    - `formula_text`: raw `<f>` content (without the leading `=`).
    - `formula_array_marker`: `<f>` `t` attribute. `'array'` / `'shared'` /
      `'dataTable'` mark non-default formula behaviors.
    - `formula_array_ref`: `<f>` `ref` attribute. For dynamic-array anchors,
      this is the SPILL RANGE (e.g. `A1:A5` for a 5-row SEQUENCE).
    - `formula_namespaces`: set of OOXML function-name prefixes appearing
      in formula_text (`_xlfn`, `_xlfn._xlws`, `_xlpm`, `_xludf`).
    """

    ref: str
    t: str | None = None
    s: int | None = None
    cm: int | None = None
    vm: int | None = None
    formula_text: str | None = None
    formula_array_marker: str | None = None
    formula_array_ref: str | None = None
    formula_namespaces: list[str] | None = None


@dataclass
class RichCell:
    """Per-cell rich data captured from openpyxl + (optionally) raw OOXML XML.

    Internal driver representation. The public driver output is serialized from
    this shape into RichCellValue JSON; the scalar reader below is retained only
    for legacy/debug callers.
    """

    value: Any  # primitive (int/float/str/bool/None) or datetime or CellRichText
    data_type: str  # OOXML t: 'n'/'s'/'str'/'b'/'e'/'d'/'inlineStr'/'f'
    number_format: str = "General"
    is_date: bool = False
    hyperlink: dict[str, Any] | None = None
    comment: dict[str, Any] | None = None
    rich_runs: list[dict[str, Any]] | None = None
    # Raw OOXML XML data, populated when a RawXmlReader is available.
    raw: RawCellData | None = None
    # Live xlwings `.api` properties captured before the workbook closed.
    surface_b: SurfaceB | None = None


_OOXML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg-rels": "http://schemas.openxmlformats.org/package/2006/relationships",
}

# Classic Excel 7-error set (PrimitiveValue kind: "error"). Anything else
# (#SPILL!, #CALC!, #FIELD!, #BUSY!, etc.) projects to kind: "extended-error"
# per D1.A.3 / coalescing doc.
_CLASSIC_ERROR_SENTINELS: frozenset[str] = frozenset({
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
})


# Modern-error integer codes from xl/richData/rdRichValueStructure.xml
# (CT_RichValueStructure with t="_error"). Per MS-XLSX §2.3.6.1.1-.10.
# MS-XLSX 29.1 claims 0-3, 5-7, 15-16 are reserved/unallocated, but empirical
# regen 2026-05-24 found Excel 2024+ emitting errorType=6 from REGEXEXTRACT-
# no-match (which displays as #N/A in the Excel UI). Spec is outdated; map
# extended empirically.
# errorType=17 is also #BUSY! (a "waiting" sub-form with `targetValue`).
_RICH_VALUE_ERROR_TYPE_MAP: dict[int, str] = {
    4: "#NAME?",
    6: "#N/A",  # empirical 2026-05-24 — Excel 2024+ REGEXEXTRACT-no-match; spec marked "reserved"
    8: "#SPILL!",
    9: "#CONNECT!",
    10: "#BLOCKED!",
    11: "#UNKNOWN!",
    12: "#FIELD!",
    13: "#CALC!",
    14: "#BUSY!",
    17: "#BUSY!",
    18: "#EXTERNAL!",
    19: "#TIMEOUT!",
}

# Captures the four OOXML function-name namespace prefixes the audit
# identified: `_xlfn.<NAME>`, `_xlfn._xlws.<NAME>` (worksheet-bound),
# `_xlpm.<param>` (LAMBDA parameter), `_xludf.<NAME>` (unknown-UDF fallback).
# Two-part prefix `_xlfn._xlws` listed first so alternation matches greedily.
_FORMULA_NAMESPACE_RE = re.compile(
    r"(_xlfn\._xlws|_xlfn|_xlpm|_xludf)\.(?:[A-Za-z][A-Za-z0-9_]*)"
)


def _resolve_rel_target(target: str, rels_dir: str) -> str:
    """Resolve an OOXML relationship Target attribute to a zip-internal path.

    OOXML rels targets can be absolute (leading `/`, from zip root) or
    relative to the directory containing the .rels file. openpyxl writes
    worksheet rels as absolute but styles/theme as relative — both forms
    appear in real files.
    """
    if target.startswith("/"):
        return target.lstrip("/")
    return rels_dir.rstrip("/") + "/" + target


class RawXmlReader:
    """Lazy reader over a saved xlsx zip, extracting OOXML fields openpyxl
    drops or normalizes away.

    Opens the zip once (caller is responsible for `close()`); pre-parses the
    workbook→sheet mapping. Per-sheet XML is parsed on first cell lookup,
    cached thereafter.

    Usage:
        with RawXmlReader(xlsx_path) as reader:
            raw = reader.get_cell("Sheet1", "A1")
            hyperlinks = reader.get_hyperlinks("Sheet1")
    """

    def __init__(self, xlsx_path: str) -> None:
        self._zip = zipfile.ZipFile(xlsx_path)
        self._sheet_paths: dict[str, str] = {}
        self._sheet_cells: dict[str, dict[str, RawCellData]] = {}
        self._sheet_hyperlinks: dict[str, list[dict[str, Any]]] = {}
        self._rich_value_index_built: bool = False
        # vm-index (1-based) -> (metadataType-name, futureMetadata-block-index)
        # populated from <valueMetadata> in xl/metadata.xml.
        self._vm_to_record: list[tuple[str, int] | None] = []
        # XLRICHVALUE futureMetadata block-index (0-based) -> rv entry index
        # (0-based) via <ext><rvb i="...">.
        self._rv_block_to_rv: list[int | None] = []
        self._rv_entries: list[ET.Element] = []
        self._rv_structures: list[ET.Element] = []
        self._build_sheet_index()

    def _read_part_case_insensitive(self, target: str) -> bytes | None:
        """Read a zip part allowing for case differences. Excel for Mac saves
        rich-value parts as lowercase (`rdrichvalue.xml`), Windows uses
        camelCase (`rdRichValue.xml`). Zip lookups are case-sensitive.
        """
        target_lower = target.lower()
        for name in self._zip.namelist():
            if name.lower() == target_lower:
                return self._zip.read(name)
        return None

    def _build_sheet_index(self) -> None:
        try:
            wb_xml = self._zip.read("xl/workbook.xml").decode("utf8")
            rels_xml = self._zip.read("xl/_rels/workbook.xml.rels").decode("utf8")
        except KeyError:
            return  # malformed xlsx; leave indexes empty
        wb_root = ET.fromstring(wb_xml)
        rels_root = ET.fromstring(rels_xml)
        rid_to_path: dict[str, str] = {}
        for rel in rels_root.findall("pkg-rels:Relationship", _OOXML_NS):
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rid:
                rid_to_path[rid] = _resolve_rel_target(target, "xl/")
        sheets_el = wb_root.find("main:sheets", _OOXML_NS)
        if sheets_el is not None:
            for sht in sheets_el.findall("main:sheet", _OOXML_NS):
                name = sht.attrib.get("name")
                rid = sht.attrib.get(f"{{{_OOXML_NS['r']}}}id")
                if name and rid and rid in rid_to_path:
                    self._sheet_paths[name] = rid_to_path[rid]

    def _parse_sheet(self, sheet_name: str) -> None:
        if sheet_name in self._sheet_cells:
            return
        path = self._sheet_paths.get(sheet_name)
        if path is None:
            self._sheet_cells[sheet_name] = {}
            self._sheet_hyperlinks[sheet_name] = []
            return
        try:
            xml = self._zip.read(path).decode("utf8")
        except KeyError:
            self._sheet_cells[sheet_name] = {}
            self._sheet_hyperlinks[sheet_name] = []
            return
        root = ET.fromstring(xml)
        cells: dict[str, RawCellData] = {}
        for row in root.iter(f"{{{_OOXML_NS['main']}}}row"):
            for c in row.findall(f"{{{_OOXML_NS['main']}}}c"):
                ref = c.attrib.get("r")
                if not ref:
                    continue
                raw = RawCellData(ref=ref)
                raw.t = c.attrib.get("t")
                s_attr = c.attrib.get("s")
                if s_attr is not None:
                    try:
                        raw.s = int(s_attr)
                    except ValueError:
                        pass
                cm_attr = c.attrib.get("cm")
                if cm_attr is not None:
                    try:
                        raw.cm = int(cm_attr)
                    except ValueError:
                        pass
                vm_attr = c.attrib.get("vm")
                if vm_attr is not None:
                    try:
                        raw.vm = int(vm_attr)
                    except ValueError:
                        pass
                f_el = c.find(f"{{{_OOXML_NS['main']}}}f")
                if f_el is not None:
                    raw.formula_text = f_el.text or ""
                    raw.formula_array_marker = f_el.attrib.get("t")
                    raw.formula_array_ref = f_el.attrib.get("ref")
                    if raw.formula_text:
                        ns_matches = _FORMULA_NAMESPACE_RE.findall(raw.formula_text)
                        if ns_matches:
                            # Preserve first-seen order, dedup.
                            seen: dict[str, None] = {}
                            for ns in ns_matches:
                                seen[ns] = None
                            raw.formula_namespaces = list(seen.keys())
                cells[ref] = raw
        self._sheet_cells[sheet_name] = cells

        # Sheet-level <hyperlinks> block + its rels file for URL resolution.
        hyperlinks_el = root.find(f"{{{_OOXML_NS['main']}}}hyperlinks")
        sheet_rels_path = path.replace("xl/worksheets/", "xl/worksheets/_rels/") + ".rels"
        rels_lookup: dict[str, str] = {}
        try:
            sheet_rels_xml = self._zip.read(sheet_rels_path).decode("utf8")
            sheet_rels_root = ET.fromstring(sheet_rels_xml)
            for rel in sheet_rels_root.findall("pkg-rels:Relationship", _OOXML_NS):
                rid = rel.attrib.get("Id", "")
                rels_lookup[rid] = rel.attrib.get("Target", "")
        except KeyError:
            pass
        hyperlinks: list[dict[str, Any]] = []
        if hyperlinks_el is not None:
            for hl in hyperlinks_el.findall(f"{{{_OOXML_NS['main']}}}hyperlink"):
                entry = dict(hl.attrib)
                rid = entry.get(f"{{{_OOXML_NS['r']}}}id")
                if rid and rid in rels_lookup:
                    entry["resolved_target"] = rels_lookup[rid]
                hyperlinks.append(entry)
        self._sheet_hyperlinks[sheet_name] = hyperlinks

    def get_cell(self, sheet_name: str, cell_ref: str) -> RawCellData | None:
        self._parse_sheet(sheet_name)
        return self._sheet_cells.get(sheet_name, {}).get(cell_ref)

    def get_hyperlinks(self, sheet_name: str) -> list[dict[str, Any]]:
        self._parse_sheet(sheet_name)
        return self._sheet_hyperlinks.get(sheet_name, [])

    def _parse_rich_value_index(self) -> None:
        # Per MS-XLSX §2.2.4.4 / ECMA-376 §18.9, vm dereferences as a chain
        # through valueMetadata + futureMetadata + rdRichValue + rdRichValueStructure:
        #   cell @vm (1-based)
        #     -> valueMetadata/bk[vm-1]/rc (t=1-based metadataType, v=0-based futureMetadata block)
        #     -> metadataTypes/metadataType[index=t]/@name (e.g. "XLRICHVALUE")
        #     -> futureMetadata[name="XLRICHVALUE"]/bk[v]/extLst/ext/rvb/@i (0-based rv index)
        #     -> rdRichValue.xml/rv[i]/@s (0-based structure index)
        #     -> rdRichValueStructure.xml/s[s_idx]
        # Lazy + idempotent + tolerant of missing parts.
        # NB: rich-value parts have inconsistent casing across Excel platforms
        # (Mac saves `rdrichvalue.xml` lowercase; Windows uses `rdRichValue.xml`).
        # Always read via _read_part_case_insensitive.
        if self._rich_value_index_built:
            return
        self._rich_value_index_built = True

        try:
            meta_xml = self._zip.read("xl/metadata.xml").decode("utf8")
        except KeyError:
            return
        meta_root = ET.fromstring(meta_xml)
        main_ns = _OOXML_NS["main"]

        # 1. Parse metadataTypes: 1-based index -> name.
        type_index_to_name: dict[int, str] = {}
        types_el = meta_root.find(f"{{{main_ns}}}metadataTypes")
        if types_el is not None:
            for i, mt in enumerate(
                types_el.findall(f"{{{main_ns}}}metadataType"), start=1
            ):
                type_index_to_name[i] = mt.attrib.get("name", "")

        # 2. Parse valueMetadata: each bk has 1+ rc records; pick the
        #    XLRICHVALUE one when present (errors use this path).
        vm_el = meta_root.find(f"{{{main_ns}}}valueMetadata")
        if vm_el is not None:
            for bk in vm_el.findall(f"{{{main_ns}}}bk"):
                chosen: tuple[str, int] | None = None
                for rc in bk.findall(f"{{{main_ns}}}rc"):
                    try:
                        t_idx = int(rc.attrib["t"])
                        v_idx = int(rc.attrib["v"])
                    except (KeyError, ValueError):
                        continue
                    type_name = type_index_to_name.get(t_idx, "")
                    if type_name == "XLRICHVALUE":
                        chosen = (type_name, v_idx)
                        break
                    if chosen is None:
                        chosen = (type_name, v_idx)
                self._vm_to_record.append(chosen)

        # 3. Parse XLRICHVALUE futureMetadata: 0-based block-index -> rv index.
        for fm in meta_root.findall(f"{{{main_ns}}}futureMetadata"):
            if fm.attrib.get("name") != "XLRICHVALUE":
                continue
            for bk in fm.findall(f"{{{main_ns}}}bk"):
                rvb = bk.find(".//{*}rvb")
                if rvb is None:
                    self._rv_block_to_rv.append(None)
                    continue
                try:
                    self._rv_block_to_rv.append(int(rvb.attrib["i"]))
                except (KeyError, ValueError):
                    self._rv_block_to_rv.append(None)
            break

        # 4. Read rdRichValue.xml and rdRichValueStructure.xml case-insensitively.
        rv_bytes = self._read_part_case_insensitive("xl/richData/rdRichValue.xml")
        if rv_bytes is None:
            return
        rv_root = ET.fromstring(rv_bytes.decode("utf8"))
        self._rv_entries = list(rv_root.findall("{*}rv"))

        struct_bytes = self._read_part_case_insensitive(
            "xl/richData/rdRichValueStructure.xml"
        )
        if struct_bytes is None:
            return
        struct_root = ET.fromstring(struct_bytes.decode("utf8"))
        self._rv_structures = list(struct_root.findall("{*}s"))

    def resolve_vm(self, vm: int) -> dict[str, Any] | None:
        """Resolve a cell's @vm attribute into a typed modern-error descriptor.

        Returns None when the rich value is not an error (e.g. Linked Data
        Type, web image), when the indirection chain is missing or
        malformed, or when the integer errorType is unrecognized.

        Returns dict with:
          - `symbol`: '#'-prefixed sentinel (e.g. '#SPILL!'); for unknown
            integer codes, the literal string `errorType=<N>`.
          - `errorType`: raw int from rdRichValueStructure.xml.
          - `subType`: raw int (opaque per MS-XLSX §2.3.6.1.3; no canonical
            integer-to-string map is published).
          - `extras`: dict of any other KVPs in the structure (`colOffset`,
            `rwOffset`, `field`, `targetValue`, etc.).

        Per ECMA-376 §18.3.1.4: the cell @vm attribute is 1-based.

        Implementation grounded in MS-XLSX spec and empirically verified for
        `#SPILL!` via probe C9. The remaining modern-error codes are
        spec-grounded until covered by additional fixtures.
        """
        self._parse_rich_value_index()

        if vm < 1 or vm > len(self._vm_to_record):
            return None
        record = self._vm_to_record[vm - 1]
        if record is None:
            return None
        type_name, block_idx = record
        if type_name != "XLRICHVALUE":
            # Could be XLDAPR (dynamic-array properties) or other; not an
            # error rich value. resolve_vm is scoped to errors only for now.
            return None
        if block_idx < 0 or block_idx >= len(self._rv_block_to_rv):
            return None
        rv_idx = self._rv_block_to_rv[block_idx]
        if rv_idx is None or rv_idx < 0 or rv_idx >= len(self._rv_entries):
            return None
        rv = self._rv_entries[rv_idx]
        s_attr = rv.attrib.get("s")
        if s_attr is None:
            return None
        try:
            s_idx = int(s_attr)
        except ValueError:
            return None
        if s_idx < 0 or s_idx >= len(self._rv_structures):
            return None
        struct = self._rv_structures[s_idx]
        if struct.attrib.get("t") != "_error":
            return None

        keys = struct.findall("{*}k")
        values = rv.findall("{*}v")
        kvp: dict[str, str] = {}
        for k_el, v_el in zip(keys, values):
            name = k_el.attrib.get("n")
            if name:
                kvp[name] = v_el.text or ""

        err_str = kvp.get("errorType")
        if err_str is None:
            return None
        try:
            err_int = int(err_str)
        except ValueError:
            return None

        result: dict[str, Any] = {
            "symbol": _RICH_VALUE_ERROR_TYPE_MAP.get(err_int, f"errorType={err_int}"),
            "errorType": err_int,
        }
        if "subType" in kvp:
            try:
                result["subType"] = int(kvp["subType"])
            except ValueError:
                result["subType"] = kvp["subType"]
        extras = {n: v for n, v in kvp.items() if n not in ("errorType", "subType")}
        if extras:
            result["extras"] = extras
        return result

    def close(self) -> None:
        self._zip.close()

    def __enter__(self) -> "RawXmlReader":
        return self

    def __exit__(self, *args: Any) -> None:
        self.close()


def _build_rich_cell(cell: Any) -> RichCell:
    """Build a RichCell from a single openpyxl Cell object."""
    rich = RichCell(
        value=cell.value,
        data_type=cell.data_type,
        number_format=cell.number_format,
        is_date=cell.is_date,
    )
    if cell.hyperlink is not None:
        rich.hyperlink = {
            "target": cell.hyperlink.target,
            "display": cell.hyperlink.display,
            "tooltip": cell.hyperlink.tooltip,
            "location": cell.hyperlink.location,
        }
    if cell.comment is not None:
        rich.comment = {
            "text": cell.comment.text,
            "author": cell.comment.author,
        }
    if isinstance(cell.value, CellRichText):
        # CellRichText is iterable: yields either str blocks (no formatting)
        # or TextBlock objects with .text + .font (InlineFont).
        runs: list[dict[str, Any]] = []
        for block in cell.value:
            if isinstance(block, str):
                runs.append({"text": block, "format": None})
            else:
                fmt: dict[str, Any] = {}
                if block.font is not None:
                    for attr in ("b", "i", "u", "strike", "color", "rFont", "sz"):
                        val = getattr(block.font, attr, None)
                        if val is not None:
                            fmt[attr] = str(val)
                runs.append({"text": block.text, "format": fmt or None})
        rich.rich_runs = runs
    return rich


def read_sheet_result_rich(
    ws: Any,
    raw_reader: RawXmlReader | None = None,
    surface_b: dict[tuple[int, int], SurfaceB] | None = None,
    top: int = TARGET_ROW,
    left: int = TARGET_COL,
    rows: int = SPILL_ROWS,
    cols: int = SPILL_COLS,
) -> list[list[RichCell | None]]:
    """Per-cell openpyxl reads producing a RichCell per populated cell,
    optionally enriched with raw OOXML XML fields and Surface B live
    `.api` properties.

    Returns the full rows × cols read window anchored at (top, left); empty cells
    are None.
    Trimming happens at the scalar-collapse boundary (`read_sheet_result`).

    Slower than the legacy `iter_rows(values_only=True)` path (~10× per the
    existing comment) due to Cell-object instantiation, but captures
    data_type, number_format, is_date, hyperlink, comment, and rich-text
    runs.

    When `raw_reader` is provided, additionally populates `rich.raw` with
    OOXML fields openpyxl drops: `cm` and `vm` metadata indexes,
    `<f t="array" ref="...">` spill markers, and namespace prefixes
    (`_xlfn.` / `_xlpm.` / `_xludf.`).

    When `surface_b` is provided (a dict keyed by 1-indexed (row, col)),
    populates `rich.surface_b` with the captured `.api.Range.*` properties.
    """
    sheet_name = ws.title
    grid: list[list[RichCell | None]] = []
    for row_idx in range(top, top + rows):
        row: list[RichCell | None] = []
        for col_idx in range(left, left + cols):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Untouched cells: data_type 'n' (default) + value None.
            # But: if the raw reader says there's a <c> element here (e.g.
            # a spill recipient with no openpyxl-visible value), still keep
            # the cell so the trim logic and raw data survive.
            raw = raw_reader.get_cell(sheet_name, cell.coordinate) if raw_reader else None
            sb = surface_b.get((row_idx, col_idx)) if surface_b else None
            if cell.value is None and cell.data_type == "n" and raw is None and sb is None:
                row.append(None)
                continue
            rich = _build_rich_cell(cell)
            if raw is not None:
                rich.raw = raw
            if sb is not None:
                rich.surface_b = sb
            row.append(rich)
        grid.append(row)
    return grid


def _serialize_value2(v: Any, epoch: datetime) -> Any:
    """Coerce a Surface B Value2 reading into a JSON-safe scalar.

    Per the RichCellValue contract: ExcelExtras.value2 is
    number|string|boolean|null. Dates convert to serials. Anything else
    (appscript Reference, COM error variant, …) returns None — capture
    already filters via _is_scalar_value but this stays defensive.
    """
    if v is None or isinstance(v, bool) or isinstance(v, (int, float, str)):
        return v
    if isinstance(v, (datetime, date, time)):
        return _dt_to_serial(v, epoch)
    return None


def _primitive_from_rich(
    rich: RichCell,
    epoch: datetime,
    raw_reader: RawXmlReader | None,
) -> dict[str, Any]:
    """Map a RichCell to a contract PrimitiveValue dict.

    Modern-error precedence: if the cell has a `vm` indirection that
    resolves to an `_error` rich value, the primitive becomes
    `kind: "extended-error"` regardless of what openpyxl returned for
    `cell.value` (which is often the bare sentinel string fallback). This
    is the first production code path that exercises `resolve_vm` — see
    D1.A.3 + D9 in the coalescing doc.

    Excel never emits `kind: "null"` (no propagatable runtime Null type);
    nullish output is always `kind: "blank"` per the β model.
    """
    if rich.raw and rich.raw.vm and raw_reader:
        descriptor = raw_reader.resolve_vm(rich.raw.vm)
        if descriptor:
            primitive: dict[str, Any] = {
                "kind": "extended-error",
                "sentinel": descriptor["symbol"],
            }
            if "errorType" in descriptor:
                primitive["error_type"] = descriptor["errorType"]
            return primitive

    v = rich.value
    if v is None:
        return {"kind": "blank", "reason": "untouched"}
    if isinstance(v, bool):
        return {"kind": "boolean", "value": v}
    if isinstance(v, (int, float)):
        if isinstance(v, float):
            if v != v or v in (float("inf"), float("-inf")):
                return {"kind": "error", "sentinel": "#NUM!"}
            if v.is_integer() and abs(v) < 2**53:
                v = int(v)
        return {"kind": "number", "value": v}
    if isinstance(v, str):
        if _ERROR_STR_RE.match(v):
            sentinel = v
            if sentinel in _CLASSIC_ERROR_SENTINELS:
                return {"kind": "error", "sentinel": sentinel}
            return {"kind": "extended-error", "sentinel": sentinel}
        return {"kind": "string", "value": v}
    if isinstance(v, (datetime, date, time)):
        serial = _dt_to_serial(v, epoch)
        if isinstance(serial, float) and serial.is_integer() and abs(serial) < 2**53:
            serial = int(serial)
        return {"kind": "number", "value": serial}
    if isinstance(v, CellRichText):
        # The collapsed string flattens runs; rich_runs is in engine extras.
        collapsed = "".join(
            block if isinstance(block, str) else block.text for block in v
        )
        return {"kind": "rich-text", "collapsed": collapsed}
    return {"kind": "string", "value": str(v)}


def _excel_extras(rich: RichCell, raw_reader: RawXmlReader | None, epoch: datetime) -> dict[str, Any]:
    """Build the ExcelExtras dict per the RichCellValue contract."""
    extras: dict[str, Any] = {"platform": "excel"}
    if rich.data_type:
        extras["data_type"] = rich.data_type
    if rich.is_date:
        extras["is_date"] = True
    if rich.comment:
        extras["comment"] = rich.comment
    if rich.rich_runs:
        extras["rich_runs"] = rich.rich_runs

    if rich.raw:
        raw_xml: dict[str, Any] = {}
        if rich.raw.t is not None:
            raw_xml["t"] = rich.raw.t
        if rich.raw.s is not None:
            raw_xml["s"] = rich.raw.s
        if rich.raw.cm is not None:
            raw_xml["cm"] = rich.raw.cm
        if rich.raw.vm is not None:
            raw_xml["vm"] = rich.raw.vm
        if rich.raw.formula_text is not None:
            raw_xml["formula_text"] = rich.raw.formula_text
        if rich.raw.formula_array_marker is not None:
            raw_xml["formula_array_marker"] = rich.raw.formula_array_marker
        if rich.raw.formula_array_ref is not None:
            raw_xml["formula_array_ref"] = rich.raw.formula_array_ref
        if rich.raw.formula_namespaces:
            raw_xml["formula_namespaces"] = rich.raw.formula_namespaces
        if raw_xml:
            extras["raw_xml"] = raw_xml

    # D9 modern-error rich descriptor (only for `_error` rich values; None
    # for Linked Data Types / web images that also use the vm indirection).
    if rich.raw and rich.raw.vm and raw_reader:
        descriptor = raw_reader.resolve_vm(rich.raw.vm)
        if descriptor:
            detail: dict[str, Any] = {"error_type": descriptor["errorType"]}
            if "subType" in descriptor:
                detail["sub_type"] = descriptor["subType"]
            if "extras" in descriptor:
                detail["extras"] = descriptor["extras"]
            extras["modern_error_detail"] = detail

    if rich.surface_b:
        # Live xlwings capture (Windows COM path)
        sb = rich.surface_b
        if sb.value2 is not None:
            extras["value2"] = _serialize_value2(sb.value2, epoch)
        if sb.display_number_format is not None:
            extras["display_format"] = {"number_format": sb.display_number_format}
        if sb.saved_as_array is not None:
            extras["saved_as_array"] = sb.saved_as_array
    elif _IS_DARWIN:
        # Mac fallback: live `.api` capture crashes Excel at scale (see
        # _capture_surface_b_for_sheet docstring). Of the three Surface B
        # fields, only `saved_as_array` adds unique signal not already
        # available via other paths:
        #   - value2: redundant with primitive value on Mac. Windows COM
        #     Value2 returns CVErr variants for error cells (vs primitive
        #     sentinel); on Mac there's no equivalent to recover, and for
        #     non-error cells value2 ≡ primitive.value (dates already
        #     serial-converted via _value_to_cell). Skip to avoid duplication
        #     and the openpyxl error-string-fallback noise (which doesn't
        #     match the Windows CVErr semantics anyway).
        #   - display_format.number_format: the post-conditional-formatting
        #     number format. KNOWN LIMITATION (decided 2026-05-30, documented
        #     in driver-surface-coalescing-2026-05-23.md "Pickup"): it diverges
        #     from the base number_format (already captured via openpyxl, see
        #     `out["number_format"]` below) ONLY when a CF rule carries a
        #     number-format dxf whose condition is true for the cell — rare, and
        #     no corpus test sets up conditional formatting, so it is redundant
        #     with number_format for every current test. On Mac the live `.api`
        #     read is additionally blocked by the AE bridge crash. Not derivable
        #     from the saved file (needs runtime CF rule evaluation). Revisit via
        #     osascript-bulk / Office.js / CF-from-XML only when a CF-number-
        #     format test actually needs it.
        #   - saved_as_array: equivalent to the persisted `<f t="array">`
        #     marker (writer-heuristic's actual decision). raw_xml carries it.
        if rich.raw and rich.raw.formula_array_marker == "array":
            extras["saved_as_array"] = True

    return extras


def _rich_cell_to_json(
    rich: RichCell | None,
    epoch: datetime,
    raw_reader: RawXmlReader | None,
) -> dict[str, Any] | None:
    """Serialize a RichCell to the canonical RichCellValue contract shape."""
    if rich is None:
        return None
    out: dict[str, Any] = {
        "primitive": _primitive_from_rich(rich, epoch, raw_reader),
        "engine": _excel_extras(rich, raw_reader, epoch),
    }
    if rich.raw and rich.raw.formula_text:
        # Strip leading "=" per RichCellValue.formula convention. OOXML
        # formula_text already has no leading "=" in most cases, but guard.
        ft = rich.raw.formula_text
        out["formula"] = ft[1:] if ft.startswith("=") else ft
    if rich.number_format and rich.number_format != "General":
        # Excel doesn't have a typed number-format enum like gsheets;
        # surface the raw format string as `pattern`.
        out["number_format"] = {"pattern": rich.number_format}
    if rich.hyperlink:
        target = rich.hyperlink.get("target") or rich.hyperlink.get("location")
        if target:
            out["hyperlink"] = target
    return out


def _trim_rich_json_grid(grid: list[list[Any]]) -> list[list[Any]]:
    """Trim trailing all-None rows/cols. Mirrors the legacy scalar trim."""
    max_cols = 0
    for row in grid:
        for c in range(len(row) - 1, -1, -1):
            if row[c] is not None:
                if c + 1 > max_cols:
                    max_cols = c + 1
                break
    if max_cols == 0:
        max_cols = 1
    grid = [r[:max_cols] for r in grid]

    max_rows = 0
    for r in range(len(grid) - 1, -1, -1):
        if any(v is not None for v in grid[r]):
            max_rows = r + 1
            break
    if max_rows == 0:
        max_rows = 1
    return grid[:max_rows]


def read_sheet_result_json(
    ws: Any,
    raw_reader: RawXmlReader | None = None,
    surface_b: dict[tuple[int, int], SurfaceB] | None = None,
    top: int = TARGET_ROW,
    left: int = TARGET_COL,
    rows: int = SPILL_ROWS,
    cols: int = SPILL_COLS,
) -> list[list[Any]]:
    """Public rich-JSON-grid output matching the RichCellValue contract.

    Each cell is `RichCellValue | None`, read from the rows × cols window anchored
    at (top, left). Trailing-null rows/cols are trimmed. This is the canonical
    driver output post-coalescing; the scalar `read_sheet_result` is retained for
    any legacy callers.
    """
    epoch = getattr(ws.parent, "epoch", None) or _DATE_EPOCH
    rich_grid = read_sheet_result_rich(
        ws, raw_reader=raw_reader, surface_b=surface_b, top=top, left=left, rows=rows, cols=cols
    )
    json_grid: list[list[Any]] = [
        [_rich_cell_to_json(rc, epoch, raw_reader) for rc in row]
        for row in rich_grid
    ]
    return _trim_rich_json_grid(json_grid)


def read_sheet_result(
    ws: Any,
    raw_reader: RawXmlReader | None = None,
    surface_b: dict[tuple[int, int], SurfaceB] | None = None,
) -> list[list[Any]]:
    """Public scalar-grid output for backward compat.

    Collapses RichCell records to canonical scalar CellValue via
    `_value_to_cell`. Trims trailing null cols/rows.

    The driver now emits `read_sheet_result_json`; this scalar path remains
    useful for probes and comparisons against the legacy fixture shape.
    """
    epoch = getattr(ws.parent, "epoch", None) or _DATE_EPOCH
    rich_grid = read_sheet_result_rich(ws, raw_reader=raw_reader, surface_b=surface_b)
    grid: list[list[Any]] = [
        [_value_to_cell(rc.value, epoch) if rc is not None else None for rc in row]
        for row in rich_grid
    ]

    # trim trailing null cols
    max_cols = 0
    for row in grid:
        for c in range(len(row) - 1, -1, -1):
            if row[c] is not None:
                if c + 1 > max_cols:
                    max_cols = c + 1
                break
    if max_cols == 0:
        max_cols = 1
    grid = [r[:max_cols] for r in grid]

    # trim trailing null rows
    max_rows = 0
    for r in range(len(grid) - 1, -1, -1):
        if any(v is not None for v in grid[r]):
            max_rows = r + 1
            break
    if max_rows == 0:
        max_rows = 1
    return grid[:max_rows]


# Max distinct host SHEETS per workbook (was: tasks per chunk — now that lump tasks
# co-tile onto shared sheets, the memory cap is sheet count, not task count). 25 =
# empirically safe with the formula2-per-cell spill path; 100 sheets × 50+
# dynamic-array formulas exhausts excel-for-mac memory during recalc. With dense
# tiling one workbook now packs far more than 25 TASKS (up to 25 hosts × tiles/host).
CHUNK_SIZE = 25

# excel-for-mac doesn't free memory between workbook closes; relaunch
# every N chunks to reclaim.
APP_RESTART_EVERY = 10

# D3 process-death recovery (design §6.2). A real crash (OOM / hard crash) kills
# the Excel process and leaves the `app` handle dead; the old code kept reusing it,
# so the bisect mislabeled up to APP_RESTART_EVERY*CHUNK_SIZE tasks as "rejected"
# before the periodic recycle. We now distinguish process-death from a live app
# merely rejecting a workbook, RELAUNCH on death, and bisect to attribute the crash
# to the one culprit (siblings re-run on the fresh app) — recorded as `crashed`.
# Real-crash detection can't be unit-verified without a destructive crash, so an
# env-gated sentinel simulates death to exercise the recovery control flow.
_CRASH_SENTINEL = "=ASSAY_FORCE_CRASH()"


class ExcelProcessDeath(Exception):
    """Excel's process died (OOM / hard crash): the app handle is dead and must be
    relaunched — distinct from a formula a *live* app rejects (caller bisects)."""


def _app_alive(app: Any) -> bool:
    # Best-effort liveness probe: a dead Excel process makes app property access
    # raise. Guards the real-crash path; the simulated-crash hook raises
    # ExcelProcessDeath directly and does not depend on this.
    try:
        _ = app.pid
        return True
    except Exception:  # noqa: BLE001
        return False


class _AppHolder:
    """Owns the current Excel app; relaunch() swaps in a fresh process on death/recycle."""

    def __init__(self) -> None:
        self._app: Any = None

    def get(self) -> Any:
        if self._app is None:
            self._app = xw.App(visible=False, add_book=False)
            _configure_app(self._app)
        return self._app

    def relaunch(self) -> Any:
        self.quit()
        return self.get()

    def quit(self) -> None:
        if self._app is not None:
            try:
                self._app.quit()
            except Exception:  # noqa: BLE001
                pass
            self._app = None


def _try_recalc(
    app: Any,
    tmp: str,
    label: str,
    tasks: list[dict[str, Any]],
) -> list[dict[str, Any]] | None:
    # one recalc attempt. returns per-sheet result list on success, or None
    # when excel rejected the workbook outright (caller bisects).
    # per-sheet read-level failures still return {"error": ...} entries.
    # on workbook rejection we delete the xlsx so excel doesn't queue it
    # for AutoRecovery (which would surface a restore-prompt later).
    if os.environ.get("ASSAY_SIMULATE_CRASH") and any(
        t.get("formula") == _CRASH_SENTINEL for t in tasks
    ):
        # test-only (env-gated): simulate a process death so the D3 recovery is
        # exercisable without a destructive real crash. Raised directly (the app
        # stays alive); the holder relaunches anyway, exercising relaunch + bisect.
        raise ExcelProcessDeath(f"{label}: simulated crash ({_CRASH_SENTINEL})")
    xlsx = os.path.join(tmp, f"{label}.xlsx")
    log(f"{label}: writing {len(tasks)} task(s)...")
    # sheet_of_task[i] = the (possibly shared) sheet task i lives on; co-tiled lumps
    # collapse onto one sheet, so sheet count ≤ task count (the amortization).
    sheet_of_task = build_workbook(tasks, xlsx)
    # (sheet_name, formula, top, left, rows, cols) per task — the TS packing plan's
    # placement, legacy AA1/20×20 when unplanned. Single-source the geometry; execute.
    placed = [
        (sheet_of_task[i], t.get("formula", ""), *_placement(t)) for i, t in enumerate(tasks)
    ]

    surface_b_by_sheet: dict[str, dict[tuple[int, int], SurfaceB]] = {}
    try:
        surface_b_by_sheet = recalc_with_excel(app, xlsx, placed)
    except Exception as e:
        try:
            os.unlink(xlsx)
        except OSError:
            pass
        if not _app_alive(app):
            # The Excel PROCESS died (not a mere workbook rejection): the handle is
            # dead. Signal D3 recovery (relaunch + bisect-attribute, upstream).
            log(f"{label}: Excel PROCESS DIED: {type(e).__name__}: {e!r}")
            raise ExcelProcessDeath(f"{label}: {type(e).__name__}: {e}") from e
        log(f"{label}: workbook rejected by Excel: {type(e).__name__}: {e!r}")
        return None

    log(f"{label}: reading cached values...")
    wb = load_workbook(xlsx, data_only=True, rich_text=True)
    # One RawXmlReader per xlsx — opens the zip once, lazily parses each
    # sheet's XML on first cell lookup. The reader's `raw` fields populate
    # rich-cell records and the final RichCellValue JSON.
    raw_reader: RawXmlReader | None = None
    try:
        raw_reader = RawXmlReader(xlsx)
    except Exception as e:  # noqa: BLE001
        log(f"{label}: RawXmlReader init failed (continuing without raw XML): {e}")
    results: list[dict[str, Any]] = []
    try:
        for i, _task in enumerate(tasks):
            sheet = sheet_of_task[i]
            if sheet not in wb.sheetnames:
                results.append({"error": f"sheet {sheet} missing after recalc"})
                continue
            try:
                _sn, _f, top, left, rows, cols = placed[i]
                sb = surface_b_by_sheet.get(sheet)
                results.append(
                    {
                        "result": read_sheet_result_json(
                            wb[sheet],
                            raw_reader=raw_reader,
                            surface_b=sb,
                            top=top,
                            left=left,
                            rows=rows,
                            cols=cols,
                        )
                    }
                )
            except Exception as e:  # noqa: BLE001
                results.append({"error": f"{type(e).__name__}: {e}"})
    finally:
        if raw_reader is not None:
            try:
                raw_reader.close()
            except Exception:  # noqa: BLE001
                pass
    wb.close()
    return results


def _run_with_bisect(
    holder: "_AppHolder",
    tmp: str,
    label: str,
    tasks: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    # Bisect to isolate a poison formula. Two failure modes, distinguished:
    #   - workbook REJECTION (live app; _try_recalc -> None): a lone rejected task
    #     gets {"error": "excel rejected formula"}.
    #   - process DEATH (ExcelProcessDeath): relaunch a FRESH app, then bisect to
    #     attribute the crash; a lone crasher gets {"crashed": "process-death"} and
    #     its siblings re-run on the fresh app — no mislabel cascade (the D3 fix).
    if not tasks:
        return []
    try:
        result = _try_recalc(holder.get(), tmp, label, tasks)
    except ExcelProcessDeath:
        holder.relaunch()
        if len(tasks) == 1:
            log(f"{label}: isolated crash-engine formula: {tasks[0].get('formula', '')!r}")
            return [{"crashed": "process-death"}]
        mid = len(tasks) // 2
        left = _run_with_bisect(holder, tmp, f"{label}L", tasks[:mid])
        right = _run_with_bisect(holder, tmp, f"{label}R", tasks[mid:])
        return left + right
    if result is not None:
        return result
    if len(tasks) == 1:
        return [{"error": "excel rejected formula (would prompt AutoRecovery)"}]
    mid = len(tasks) // 2
    left = _run_with_bisect(holder, tmp, f"{label}L", tasks[:mid])
    right = _run_with_bisect(holder, tmp, f"{label}R", tasks[mid:])
    return left + right


def _excel_sandbox_tmpdir() -> str:
    override = os.environ.get("ASSAY_EXCEL_TMPDIR")
    if override:
        root = os.path.abspath(os.path.expanduser(override))
        os.makedirs(root, exist_ok=True)
        return tempfile.mkdtemp(prefix="run-", dir=root)

    # excel-for-mac is sandboxed; opening files outside its Container dir
    # triggers a "grant access" prompt per file and sometimes OSERROR -50.
    # writing temp xlsxes inside ~/Library/Containers/com.microsoft.Excel/Data
    # bypasses both. falls back to system tmp on linux/windows or no-excel.
    # see https://stackoverflow.com/a/79766529.
    candidate = os.path.expanduser("~/Library/Containers/com.microsoft.Excel/Data")
    if os.path.isdir(candidate) and os.access(candidate, os.W_OK):
        root = os.path.join(candidate, "assay-tmp")
        try:
            os.makedirs(root, exist_ok=True)
            return tempfile.mkdtemp(prefix="run-", dir=root)
        except PermissionError as e:
            log(f"Excel sandbox tmpdir unavailable: {e}; falling back to system tmp")
    return tempfile.mkdtemp(prefix="assay-excel-")


def _is_spill(entry: dict[str, Any]) -> bool:
    """True iff a result entry's top-left cell is a #SPILL!. For a CO-TILED lump that
    is an ARTIFACT — a neighbouring tile blocked the spill, not the real answer (the
    formula is reference-free, so ALONE it spills freely). Re-running it isolated
    recovers the true (window-clipped) result."""
    res = entry.get("result")
    if not res or not res[0]:
        return False
    cell = res[0][0]
    return isinstance(cell, dict) and cell.get("primitive", {}).get("sentinel") == "#SPILL!"


def _chunk_by_hosts(tasks: list[dict[str, Any]], max_hosts: int) -> list[tuple[int, int]]:
    """Slice `tasks` into contiguous [lo, hi) chunks each spanning ≤ max_hosts distinct
    host SHEETS (the per-workbook memory cap). Co-tiled lumps share a host, so a chunk
    packs far more TASKS than sheets — the amortization. A host straddling a chunk
    boundary is harmless (each tile is self-contained); it only loses a little tiling."""
    chunks: list[tuple[int, int]] = []
    lo = 0
    cur_hosts: set[tuple[str, int]] = set()
    for i, t in enumerate(tasks):
        key = _host_key(t, i)
        if key not in cur_hosts and len(cur_hosts) >= max_hosts:
            chunks.append((lo, i))
            lo = i
            cur_hosts = set()
        cur_hosts.add(key)
    if lo < len(tasks):
        chunks.append((lo, len(tasks)))
    return chunks


def run_batch(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not tasks:
        return []

    # short-circuit skip:-tagged tasks before touching excel. those formulas
    # would crash excel's open path (SPLIT, QUERY, ARRAYFORMULA, …) or
    # require external I/O (IMPORT*, GOOGLE*). emit #N/A + `skipped` meta
    # so fixtures distinguish "engine can't" from "engine errored".
    results: list[dict[str, Any] | None] = [None] * len(tasks)
    excel_indices: list[int] = []
    excel_tasks: list[dict[str, Any]] = []
    for i, t in enumerate(tasks):
        reason = t.get("skip")
        if reason:
            # Skip placeholder in the rich-JSON shape (post-coalescing).
            results[i] = {
                "result": [[{
                    "primitive": {"kind": "error", "sentinel": "#N/A"},
                    "engine": {"platform": "excel"},
                }]],
                "skipped": reason,
            }
        else:
            excel_indices.append(i)
            excel_tasks.append(t)

    if not excel_tasks:
        return [r or {"error": "unreachable"} for r in results]

    # Pack the excel-bound tasks into workbooks bounded by host SHEET count (lump
    # tasks co-tile, so far more tasks than sheets per workbook). Each workbook runs
    # under one app, bisect-on-crash/rejection to isolate poison formulas.
    tmp = _excel_sandbox_tmpdir()
    skipped_count = len(tasks) - len(excel_tasks)
    chunks = _chunk_by_hosts(excel_tasks, CHUNK_SIZE)
    stats = {"tasks": len(excel_tasks), "sheets": 0, "workbooks": len(chunks), "spill_reruns": 0}
    try:
        log(
            f"launching Excel ({len(excel_tasks)} task(s) → {len(chunks)} workbook(s), "
            f"≤{CHUNK_SIZE} sheets each, {skipped_count} pre-skipped)..."
        )
        log(f"  tmpdir: {tmp}")
        holder = _AppHolder()
        try:
            chunks_since_restart = 0
            for ci, (lo, hi) in enumerate(chunks):
                chunk = excel_tasks[lo:hi]
                # recycle excel every APP_RESTART_EVERY workbooks — long runs
                # crawl then OOM otherwise (no memory release on workbook close)
                if chunks_since_restart >= APP_RESTART_EVERY:
                    log(f"  recycling Excel app after {chunks_since_restart} workbook(s)")
                    holder.relaunch()
                    chunks_since_restart = 0
                label = f"chunk{ci}"
                chunk_results = _run_with_bisect(holder, tmp, label, chunk)
                chunks_since_restart += 1

                host_keys = [_host_key(t, j) for j, t in enumerate(chunk)]
                stats["sheets"] += len(set(host_keys))

                # #SPILL!-artifact recovery: a co-tiled lump (host shared by >1 task)
                # that came back #SPILL! was blocked by a neighbour tile, not really a
                # spill error. Re-run those ALONE (placement stripped → solo AA1 full
                # window) so they spill freely, and override.
                host_count: dict[tuple[str, int], int] = {}
                for key in host_keys:
                    host_count[key] = host_count.get(key, 0) + 1
                suspects = [
                    j
                    for j in range(len(chunk))
                    if host_count[host_keys[j]] > 1 and _is_spill(chunk_results[j])
                ]
                if suspects:
                    log(f"  {label}: {len(suspects)} co-tiled #SPILL! → isolated re-run")
                    iso_tasks = []
                    for j in suspects:
                        t = dict(chunk[j])
                        t.pop("placement", None)  # → solo sheet, legacy AA1/20×20 default window
                        iso_tasks.append(t)
                    # ONE recovery pass: the re-run is solo (placement stripped), so it
                    # can't be co-tiled and won't artifact-#SPILL! again. If a re-run STILL
                    # returns #SPILL! it's a GENUINE spill (recorded as-is) — not re-rescued.
                    iso_results = _run_with_bisect(holder, tmp, f"{label}-iso", iso_tasks)
                    for k, j in enumerate(suspects):
                        chunk_results[j] = iso_results[k]
                    stats["spill_reruns"] += len(suspects)

                for j, r in enumerate(chunk_results):
                    results[excel_indices[lo + j]] = r
        finally:
            holder.quit()
    finally:
        try:
            import shutil
            shutil.rmtree(tmp, ignore_errors=True)
        except Exception:
            pass

    stats_path = os.environ.get("ASSAY_EXCEL_STATS")
    if stats_path:
        try:
            with open(stats_path, "w", encoding="utf8") as f:
                json.dump(stats, f)
        except OSError:
            pass

    return [r if r is not None else {"error": "unreachable"} for r in results]


def main() -> int:
    argv = sys.argv[1:]

    if argv == ["--quit"]:
        try:
            for app in xw.apps:
                app.quit()
        except Exception:
            pass
        return 0

    if argv == ["--version"]:
        # opens excel briefly (~3-5s); only invoked from `assay history --record`
        app = xw.App(visible=False, add_book=False)
        try:
            print(app.version)
        finally:
            app.quit()
        return 0

    # --workbook ignored by the bulk driver — charter-style custom-package
    # workbooks aren't supported here (we always build a fresh xlsx). flag
    # accepted for cli parity; warn if passed.
    if "--workbook" in argv:
        idx = argv.index("--workbook")
        print(
            "[excel] warning: --workbook is not supported by the bulk driver; "
            "custom-package workbooks will not be loaded. Tests will run against "
            "a fresh workbook.",
            file=sys.stderr,
        )
        argv = argv[:idx] + argv[idx + 2:]

    batch = False
    if argv and argv[0] == "--batch":
        batch = True
        argv = argv[1:]

    if len(argv) != 2:
        print(
            "usage: excel_driver.py [--workbook <path>] [--batch] <in.json> <out.json>",
            file=sys.stderr,
        )
        return 2

    in_path, out_path = argv
    with open(in_path, encoding="utf8") as f:
        data = json.load(f)

    if batch:
        results = run_batch(data)
    else:
        results_list = run_batch([data])
        results = results_list[0]

    with open(out_path, "w", encoding="utf8") as f:
        json.dump(results, f)
    return 0


if __name__ == "__main__":
    sys.exit(main())
