"""libreoffice calc driver — `soffice --headless` recalcs an xlsx.

openpyxl writes the input xlsx, soffice recalcs + re-saves, openpyxl reads
cached values back via `data_only=True`. Recalc-on-load is NOT on by default
(headless treats the "prompt" default as never) so we pre-seed a fresh profile
forcing it — see `_seed_recalc_profile`; without it every formula reads blank.

requires soffice on PATH (or the standard macos location); install via
`brew install --cask libreoffice` or `apt install libreoffice-calc`.

soffice startup is ~2s so only batch mode is practical; single-task mode
exists for api parity.

protocol matches the other python drivers:
  single: [task.json, result.json]
  batch:  --batch [tasks.json, results.json]
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


_DATE_EPOCH = datetime(1899, 12, 30)

# openpyxl writes formula strings with no cached result. LibreOffice only fills
# those results in if it recalculates while loading the xlsx — and the default
# "Recalculation on File Load" mode for OOXML/ODF documents is "prompt", which in
# headless mode is a silent no-op. Without this, every formula cell reads back as
# an empty cached value (None) via data_only=True. Pre-seed a fresh user profile
# forcing "always recalculate on load" (mode 0) so `--convert-to` computes results
# before re-saving. Modes: 0=always, 1=never, 2=prompt.
_RECALC_PROFILE_XCU = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<oor:items '
    'xmlns:oor="http://openoffice.org/2001/registry" '
    'xmlns:xs="http://www.w3.org/2001/XMLSchema" '
    'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
    ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
    ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
    '</oor:items>\n'
)


def _seed_recalc_profile(profile_dir: str) -> str:
    """Create a LibreOffice user profile forcing recalc-on-load. Returns a
    file:// UserInstallation URL usable as `-env:UserInstallation=`."""
    user_dir = os.path.join(profile_dir, "user")
    os.makedirs(user_dir, exist_ok=True)
    with open(os.path.join(user_dir, "registrymodifications.xcu"), "w", encoding="utf8") as f:
        f.write(_RECALC_PROFILE_XCU)
    return Path(profile_dir).as_uri()


def _dt_to_serial(v: Any) -> float:
    if isinstance(v, datetime):
        return (v - _DATE_EPOCH).total_seconds() / 86400.0
    if isinstance(v, date):
        return float((datetime(v.year, v.month, v.day) - _DATE_EPOCH).days)
    if isinstance(v, time):
        return (v.hour * 3600 + v.minute * 60 + v.second + v.microsecond / 1e6) / 86400.0
    raise TypeError(type(v))


SOFFICE_CANDIDATES = [
    "soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
    "/usr/local/bin/soffice",
]

TARGET_ROW = 1  # AA1 — far from user grid regions
TARGET_COL = 27
SPILL_ROWS = 20
SPILL_COLS = 20

_CELL_REF_RE = re.compile(r"^([A-Z]+)(\d+)$")


def find_soffice() -> str:
    for c in SOFFICE_CANDIDATES:
        path = shutil.which(c) if "/" not in c else (c if os.path.exists(c) else None)
        if path:
            return path
    raise RuntimeError(
        "LibreOffice 'soffice' not found. Install via "
        "`brew install --cask libreoffice` or `apt install libreoffice-calc`."
    )


def parse_cell_ref(ref: str) -> tuple[int, int]:
    m = _CELL_REF_RE.match(ref.upper())
    if not m:
        raise ValueError(f"Invalid cell reference: {ref}")
    letters, row = m.group(1), int(m.group(2))
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


def write_input_workbook(tasks: list[dict[str, Any]], path: str) -> None:
    # one sheet per task t0 … tN; formula at AA1, grid at A1-style refs
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    for i, task in enumerate(tasks):
        ws = wb.create_sheet(title=f"t{i}")

        for ref, val in (task.get("grid") or {}).items():
            row, col = parse_cell_ref(ref)
            if val is None:
                continue
            if isinstance(val, dict) and "error" in val:
                ws.cell(row=row, column=col).value = val["error"]
            else:
                ws.cell(row=row, column=col).value = val

        ws.cell(row=TARGET_ROW, column=TARGET_COL).value = task.get("formula", "")

    wb.save(path)


def recalc_with_libreoffice(input_xlsx: str, out_dir: str) -> str:
    soffice = find_soffice()
    env = os.environ.copy()
    env["HOME"] = out_dir  # don't pollute the user's libreoffice profile
    # Deterministic, platform-independent profile location (the HOME-based default
    # path differs by OS). Seed it with the recalc-on-load setting so soffice
    # computes formula results during --convert-to instead of shipping blanks.
    profile_dir = os.path.join(out_dir, "lo-profile")
    user_installation = _seed_recalc_profile(profile_dir)
    cmd = [
        soffice,
        f"-env:UserInstallation={user_installation}",
        "--headless",
        "--calc",
        "--norestore",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to",
        "xlsx",
        "--outdir",
        out_dir,
        input_xlsx,
    ]
    result = subprocess.run(
        cmd, env=env, capture_output=True, text=True, timeout=600
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"soffice failed (rc={result.returncode}):\n{result.stderr}"
        )
    base = os.path.basename(input_xlsx)
    out_path = os.path.join(out_dir, base)
    if not os.path.exists(out_path):
        raise RuntimeError(f"LibreOffice did not produce output at {out_path}")
    return out_path


def cell_to_value(cell: Any) -> Any:
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        if v.startswith("#") and "!" in v:
            return {"error": v}
        # some LO builds emit error-like strings without `!` (e.g. "#NAME?")
        if re.match(r"^#[A-Z0-9/]+\??$", v):
            return {"error": v}
        return v
    if isinstance(v, (datetime, date, time)):
        serial = _dt_to_serial(v)
        if isinstance(serial, float) and serial.is_integer() and abs(serial) < 2**53:
            return int(serial)
        return serial
    return str(v)


def read_result_grid(ws: Any) -> list[list[Any]]:
    grid: list[list[Any]] = []
    for dr in range(SPILL_ROWS):
        row: list[Any] = []
        for dc in range(SPILL_COLS):
            cell = ws.cell(row=TARGET_ROW + dr, column=TARGET_COL + dc)
            row.append(cell_to_value(cell))
        grid.append(row)

    # trim trailing null cols
    max_cols = 0
    for row in grid:
        for c in range(len(row) - 1, -1, -1):
            if row[c] is not None:
                if c + 1 > max_cols:
                    max_cols = c + 1
                break
    if max_cols == 0:
        max_cols = 1
    grid = [row[:max_cols] for row in grid]

    # trim trailing null rows
    max_rows = 0
    for r in range(len(grid) - 1, -1, -1):
        if any(v is not None for v in grid[r]):
            max_rows = r + 1
            break
    if max_rows == 0:
        max_rows = 1
    return grid[:max_rows]


def run_batch(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    with tempfile.TemporaryDirectory(prefix="assay-lo-") as tmp:
        input_xlsx = os.path.join(tmp, "in.xlsx")
        write_input_workbook(tasks, input_xlsx)
        out_xlsx = recalc_with_libreoffice(input_xlsx, tmp)
        wb = load_workbook(out_xlsx, data_only=True)

        results: list[dict[str, Any]] = []
        for i in range(len(tasks)):
            sheet_name = f"t{i}"
            if sheet_name not in wb.sheetnames:
                results.append({"error": f"sheet {sheet_name} missing after recalc"})
                continue
            try:
                results.append({"result": read_result_grid(wb[sheet_name])})
            except Exception as e:  # noqa: BLE001
                results.append({"error": f"{type(e).__name__}: {e}"})
        return results


def main() -> int:
    args = sys.argv[1:]
    if args and args[0] == "--version":
        import subprocess
        try:
            out = subprocess.run(
                [find_soffice(), "--version"], capture_output=True, text=True, timeout=10,
            ).stdout.strip()
            # "LibreOffice 25.2.4.1 abc123 …" → "25.2.4.1"
            parts = out.split()
            print(parts[1] if len(parts) >= 2 else out)
        except Exception:
            print("")
        return 0
    batch = False
    if args and args[0] == "--batch":
        batch = True
        args = args[1:]
    if len(args) != 2:
        print("usage: libreoffice_driver.py [--version | [--batch] <in.json> <out.json>]", file=sys.stderr)
        return 2

    in_path, out_path = args
    with open(in_path, encoding="utf8") as f:
        data = json.load(f)

    if batch:
        results = run_batch(data)
    else:
        results_list = run_batch([data])
        results = results_list[0]

    with open(out_path, "w", encoding="utf8") as f:
        json.dump(results, f)
    return 0


if __name__ == "__main__":
    sys.exit(main())
