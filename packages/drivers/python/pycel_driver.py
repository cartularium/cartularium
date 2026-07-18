"""pycel driver — compiles an xlsx into an evaluable graph and reads cells.

protocol matches the other python drivers:
  single: [task.json, result.json]
  batch:  --batch [tasks.json, results.json]
"""

from __future__ import annotations

import json
import os
import re
import sys
import tempfile
import warnings
from typing import Any

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", category=SyntaxWarning)

import logging  # noqa: E402

# pycel logs a full traceback for every unsupported function; silence
logging.getLogger("pycel").setLevel(logging.CRITICAL)
logging.getLogger("pycel.excelformula").setLevel(logging.CRITICAL)
logging.getLogger("pycel.excelcompiler").setLevel(logging.CRITICAL)

from openpyxl import Workbook  # noqa: E402
from pycel import ExcelCompiler  # noqa: E402


# AA1 — cols A-Z reserved for grid: cells
TARGET_ROW = 1
TARGET_COL = 27
SPILL_ROWS = 20
SPILL_COLS = 20

_CELL_REF_RE = re.compile(r"^([A-Z]+)(\d+)$")
_ERROR_RE = re.compile(r"^#[A-Z0-9/]+!?\??$")


def parse_cell_ref(ref: str) -> tuple[int, int]:
    m = _CELL_REF_RE.match(ref.upper())
    if not m:
        raise ValueError(f"Invalid cell reference: {ref}")
    letters, row = m.group(1), int(m.group(2))
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


def col_letters(col: int) -> str:
    out = ""
    while col > 0:
        col, r = divmod(col - 1, 26)
        out = chr(65 + r) + out
    return out


def build_workbook(tasks: list[dict[str, Any]], path: str) -> list[str]:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    sheets: list[str] = []
    for i, task in enumerate(tasks):
        name = f"t{i}"
        sheets.append(name)
        ws = wb.create_sheet(title=name)
        for ref, val in (task.get("grid") or {}).items():
            row, col = parse_cell_ref(ref)
            if val is None:
                continue
            if isinstance(val, dict) and "error" in val:
                ws.cell(row=row, column=col).value = val["error"]
            else:
                ws.cell(row=row, column=col).value = val
        ws.cell(row=TARGET_ROW, column=TARGET_COL).value = task.get("formula", "")
    wb.save(path)
    return sheets


def unwrap(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v != v or v in (float("inf"), float("-inf")):
            return {"error": "#NUM!"}
        return int(v) if v.is_integer() and abs(v) < 2**53 else v
    if isinstance(v, str):
        if v == "":
            return None
        if _ERROR_RE.match(v):
            return {"error": v}
        return v
    if isinstance(v, (list, tuple)):
        # pycel sometimes returns nested tuples for ranges
        if v and isinstance(v[0], (list, tuple)):
            return [[unwrap(x) for x in row] for row in v]
        return [unwrap(x) for x in v]
    return str(v)


def read_task_result(xl: ExcelCompiler, sheet: str) -> list[list[Any]]:
    grid: list[list[Any]] = []
    for dr in range(SPILL_ROWS):
        row: list[Any] = []
        for dc in range(SPILL_COLS):
            ref = f"{sheet}!{col_letters(TARGET_COL + dc)}{TARGET_ROW + dr}"
            try:
                val: Any = xl.evaluate(ref)
            except Exception:  # noqa: BLE001 — unsupported function, ref error, etc.
                # only the target cell's failure is meaningful; the rest stays null
                val = {"error": "#NAME?"} if (dr == 0 and dc == 0) else None
            unwrapped = unwrap(val) if not (isinstance(val, dict) and "error" in val) else val
            # target returned a 2D block — expand into the grid
            if dr == 0 and dc == 0 and isinstance(unwrapped, list) and unwrapped and isinstance(unwrapped[0], list):
                return unwrapped  # type: ignore[return-value]
            row.append(unwrapped if not isinstance(unwrapped, list) else None)
        grid.append(row)

    # trim trailing null cols
    max_cols = 0
    for r in grid:
        for c in range(len(r) - 1, -1, -1):
            if r[c] is not None:
                if c + 1 > max_cols:
                    max_cols = c + 1
                break
    if max_cols == 0:
        max_cols = 1
    grid = [r[:max_cols] for r in grid]

    # trim trailing null rows
    max_rows = 0
    for r in range(len(grid) - 1, -1, -1):
        if any(v is not None for v in grid[r]):
            max_rows = r + 1
            break
    if max_rows == 0:
        max_rows = 1
    return grid[:max_rows]


CHUNK_SIZE = 32  # cross-sheet contamination past ~50 sheets per book


def _run_chunk(tmp: str, chunk_idx: int, tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    path = os.path.join(tmp, f"chunk{chunk_idx}.xlsx")
    sheets = build_workbook(tasks, path)

    try:
        xl = ExcelCompiler(filename=path)
    except Exception as e:  # noqa: BLE001
        return [{"error": f"{type(e).__name__}: {e}"} for _ in tasks]

    out: list[dict[str, Any]] = []
    for sheet in sheets:
        try:
            out.append({"result": read_task_result(xl, sheet)})
        except Exception as e:  # noqa: BLE001
            out.append({"error": f"{type(e).__name__}: {e}"})
    return out


def run_batch(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # chunked because pycel miscompiles cells in later sheets past ~50 per workbook;
    # each chunk gets its own ExcelCompiler.
    with tempfile.TemporaryDirectory(prefix="assay-pycel-") as tmp:
        out: list[dict[str, Any]] = []
        for i in range(0, len(tasks), CHUNK_SIZE):
            out.extend(_run_chunk(tmp, i // CHUNK_SIZE, tasks[i : i + CHUNK_SIZE]))
        return out


def main() -> int:
    args = sys.argv[1:]
    if args and args[0] == "--version":
        from importlib.metadata import PackageNotFoundError, version
        try:
            print(version("pycel"))
        except PackageNotFoundError:
            print("")
        return 0
    batch = False
    if args and args[0] == "--batch":
        batch = True
        args = args[1:]
    if len(args) != 2:
        print("usage: pycel_driver.py [--version | [--batch] <in.json> <out.json>]", file=sys.stderr)
        return 2

    in_path, out_path = args
    with open(in_path, encoding="utf8") as f:
        data = json.load(f)

    if batch:
        results = run_batch(data)
    else:
        results_list = run_batch([data])
        results = results_list[0]

    with open(out_path, "w", encoding="utf8") as f:
        json.dump(results, f)
    return 0


if __name__ == "__main__":
    sys.exit(main())
