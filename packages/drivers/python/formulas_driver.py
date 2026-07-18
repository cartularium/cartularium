"""formulas driver — pure-python `formulas` xlsx evaluator.

stages tasks as an xlsx (sheet per task, formula at AA1, grid values around it),
calcs the whole workbook, reads results from the solution dict.

protocol matches the other python drivers:
  single: [task.json, result.json]
  batch:  --batch [tasks.json, results.json]
"""

from __future__ import annotations

import json
import os
import re
import sys
import tempfile
import warnings
from typing import Any

warnings.filterwarnings("ignore", category=SyntaxWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

import formulas  # noqa: E402
from formulas import XlError  # noqa: E402
from openpyxl import Workbook  # noqa: E402
import numpy as np  # noqa: E402


TARGET_ROW = 1  # AA1 — far from user grid regions
TARGET_COL = 27
SPILL_ROWS = 20
SPILL_COLS = 20

_CELL_REF_RE = re.compile(r"^([A-Z]+)(\d+)$")


def parse_cell_ref(ref: str) -> tuple[int, int]:
    m = _CELL_REF_RE.match(ref.upper())
    if not m:
        raise ValueError(f"Invalid cell reference: {ref}")
    letters, row = m.group(1), int(m.group(2))
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


def col_letters(col: int) -> str:
    out = ""
    while col > 0:
        col, r = divmod(col - 1, 26)
        out = chr(65 + r) + out
    return out


def build_input_xlsx(tasks: list[dict[str, Any]], path: str) -> list[str]:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    sheet_names: list[str] = []
    for i, task in enumerate(tasks):
        name = f"t{i}"
        sheet_names.append(name)
        ws = wb.create_sheet(title=name)
        for ref, val in (task.get("grid") or {}).items():
            row, col = parse_cell_ref(ref)
            if val is None:
                continue
            if isinstance(val, dict) and "error" in val:
                ws.cell(row=row, column=col).value = val["error"]
            else:
                ws.cell(row=row, column=col).value = val
        ws.cell(row=TARGET_ROW, column=TARGET_COL).value = task.get("formula", "")
    wb.save(path)
    return sheet_names


def cell_key(book: str, sheet: str, row: int, col: int) -> str:
    # `formulas` uses keys like '[in.xlsx]T0'!AA1 with uppercase sheet
    return f"'[{book}]{sheet.upper()}'!{col_letters(col)}{row}"


def unwrap(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, XlError):
        return {"error": str(val)}
    if isinstance(val, (bool, np.bool_)):
        return bool(val)
    if isinstance(val, (int, np.integer)):
        return int(val)
    if isinstance(val, (float, np.floating)):
        f = float(val)
        if np.isnan(f) or np.isinf(f):
            return {"error": "#NUM!"}
        return int(f) if f.is_integer() and abs(f) < 2**53 else f
    if isinstance(val, bytes):
        val = val.decode()
    if isinstance(val, str):
        if val == "":
            return None
        # errors can leak as strings when a formula returned a cached value
        if re.match(r"^#[A-Z0-9/]+!?\??$", val):
            return {"error": val}
        return val
    if hasattr(val, "tolist"):
        return unwrap_grid(val.tolist())  # numpy / Ranges, shouldn't hit for scalars
    return str(val)


def unwrap_grid(grid: Any) -> list[list[Any]]:
    if not isinstance(grid, list):
        return [[unwrap(grid)]]
    if not grid or not isinstance(grid[0], list):
        return [[unwrap(v) for v in grid]]
    return [[unwrap(v) for v in row] for row in grid]


def read_result(sol: Any, book: str, sheet: str) -> list[list[Any]]:
    # `formulas` casts every result into the shape of its target cell, so an
    # N×M array written to the single-cell AA1 collapses to its top-left.
    # to recover the full array we pull the raw pre-cast value from the
    # workflow node's `solution_filters` — first filter = result before truncation.
    grid: list[list[Any]] = [[None] * SPILL_COLS for _ in range(SPILL_ROWS)]
    origin = cell_key(book, sheet, TARGET_ROW, TARGET_COL)
    node = sol.workflow.nodes.get(origin, {})
    filters = node.get("solution_filters") or []
    raw = filters[0] if filters else None
    if raw is None:
        entry = sol.get(origin)
        raw = entry.value if hasattr(entry, "value") else entry

    shape = getattr(raw, "shape", None)
    if shape is not None and len(shape) == 2 and shape != (1, 1):
        rows, cols = shape
        for sr in range(min(rows, SPILL_ROWS)):
            for sc in range(min(cols, SPILL_COLS)):
                grid[sr][sc] = unwrap(raw[sr][sc])
    elif shape == (1, 1):
        grid[0][0] = unwrap(raw[0][0])
    elif shape == ():
        grid[0][0] = unwrap(raw.item())
    else:
        grid[0][0] = unwrap(raw)

    # trim trailing null cols
    max_cols = 0
    for row in grid:
        for c in range(len(row) - 1, -1, -1):
            if row[c] is not None:
                if c + 1 > max_cols:
                    max_cols = c + 1
                break
    if max_cols == 0:
        max_cols = 1
    grid = [row[:max_cols] for row in grid]

    # trim trailing null rows
    max_rows = 0
    for r in range(len(grid) - 1, -1, -1):
        if any(v is not None for v in grid[r]):
            max_rows = r + 1
            break
    if max_rows == 0:
        max_rows = 1
    return grid[:max_rows]


def evaluate_one(tmp_dir: str, idx: int, task: dict[str, Any]) -> dict[str, Any]:
    # one xlsx per task so a parse error in one doesn't cascade
    path = os.path.join(tmp_dir, f"t{idx}.xlsx")
    try:
        build_input_xlsx([task], path)
        xl = formulas.ExcelModel().loads(path).finish()
        sol = xl.calculate()
        book = os.path.basename(path)
        return {"result": read_result(sol, book, "t0")}
    except Exception as e:  # noqa: BLE001
        return {"error": f"{type(e).__name__}: {e}"}
    finally:
        if os.path.exists(path):
            os.unlink(path)


def run_batch(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # `formulas` aborts the whole workbook on any parse error — so one xlsx per
    # task. all tasks share a python process to avoid Node↔Python spawn overhead.
    with tempfile.TemporaryDirectory(prefix="assay-formulas-") as tmp:
        return [evaluate_one(tmp, i, task) for i, task in enumerate(tasks)]


def main() -> int:
    args = sys.argv[1:]
    if args and args[0] == "--version":
        from importlib.metadata import PackageNotFoundError, version
        try:
            print(version("formulas"))
        except PackageNotFoundError:
            print(getattr(formulas, "__version__", "") or "")
        return 0
    batch = False
    if args and args[0] == "--batch":
        batch = True
        args = args[1:]
    if len(args) != 2:
        print("usage: formulas_driver.py [--version | [--batch] <in.json> <out.json>]", file=sys.stderr)
        return 2

    in_path, out_path = args
    with open(in_path, encoding="utf8") as f:
        data = json.load(f)

    if batch:
        results = run_batch(data)
    else:
        results_list = run_batch([data])
        results = results_list[0]

    with open(out_path, "w", encoding="utf8") as f:
        json.dump(results, f)
    return 0


if __name__ == "__main__":
    sys.exit(main())
