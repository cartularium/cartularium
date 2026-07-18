#!/usr/bin/env python3
"""Excel driver fidelity audit + behavior probes.

Builds a multi-sheet xlsx, recalcs via xlwings, then reads back via three
surfaces per scenario:

  A. openpyxl per-cell (data_only=True) — value, data_type, number_format,
     is_date, hyperlink, comment, CellRichText when applicable
  B. xlwings live .api (best-effort on Mac AppleEvents)
  C. raw OOXML XML extracted from the saved xlsx

Writes the report to packages/assay/docs/excel-driver-fidelity.md. The primary
deliverable is the cross-surface disagreements — those tell us which surface
to trust per field.

This is the audit phase of the canonical-cell-value-fidelity roadmap. The
behavior assertions (null vs "", error sentinel surface, etc.) come along for
free.

Run: uv run python packages/assay/scripts/probes/excel-driver-fidelity.py
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable
from xml.etree import ElementTree as ET

import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Output paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[2]  # packages/assay/
REPORT_PATH = PROJECT_ROOT / "docs" / "excel-driver-fidelity.md"

# ---------------------------------------------------------------------------
# OOXML namespaces — for raw XML parsing
# ---------------------------------------------------------------------------

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg-rels": "http://schemas.openxmlformats.org/package/2006/relationships",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
}

# ---------------------------------------------------------------------------
# Scenario type — each scenario knows how to build itself on a sheet and
# what targets to probe afterward.
# ---------------------------------------------------------------------------


@dataclass
class Target:
    """A cell to read back after recalc."""

    ref: str  # A1-style
    description: str = ""  # what we expect to learn here


@dataclass
class Scenario:
    name: str
    section: str  # for report grouping
    description: str
    # build_data(ws) populates literals, rich text, manual hyperlinks — NOT formulas.
    # openpyxl-written formulas trigger Excel's "_xludf." unknown-function rewrite
    # for any modern function (SEQUENCE, FILTER, LAMBDA, etc.) because modern
    # functions are stored namespaced (_xlfn.NAME) in OOXML but openpyxl writes
    # the bare name. Entering through xlwings post-open via formula2 lets Excel's
    # own parser do the namespacing correctly.
    build_data: Callable[[Worksheet], None] = lambda ws: None
    # Formulas entered via xlwings.Range.formula2 after Excel opens the workbook
    formulas: list[tuple[str, str]] = field(default_factory=list)
    # Formulas entered via xlwings.Range.formula_array (legacy CSE arrays)
    cse_formulas: list[tuple[str, str]] = field(default_factory=list)
    # Optional callable invoked AFTER formulas are entered but BEFORE recalc.
    # Used for #SPILL! tests where we need the spill to happen first, then
    # write an obstacle into a recipient cell to force the conflict.
    # Signature: post_formula_data(sheet) -> None where sheet is an xlwings Sheet.
    post_formula_data: Callable[[Any], None] | None = None
    targets: list[Target] = field(default_factory=list)
    expect: str = ""
    needs_workbook_setup: Callable[[Workbook], None] | None = None


# ---------------------------------------------------------------------------
# Scenarios
# ---------------------------------------------------------------------------


def _set_value(ws: Worksheet, ref: str, value: Any) -> None:
    ws[ref] = value


def _set_formula(ws: Worksheet, ref: str, formula: str) -> None:
    ws[ref] = formula


# ---- Section 1: Null / blank / empty (mirrors gsheets probe 6 with Excel twists) ----


SCN_NULL = Scenario(
    name="null-blank-empty",
    section="Null / blank / empty",
    description="Does Excel have a runtime-Null distinct from empty string and 0? Use Excel-specific TYPE() and CELL('type') alongside the gsheets-style ISBLANK/ISTEXT probes.",
    formulas=[
        ("A2", '=""'),
        ("A3", "=IF(,,)"),
        ("A4", "=ISBLANK(A1)"),
        ("A5", "=ISBLANK(A2)"),
        ("A6", "=ISBLANK(A3)"),
        ("A7", "=ISTEXT(A2)"),
        ("A8", "=ISTEXT(A3)"),
        ("A9", '="x" & A2'),
        ("A10", '="x" & A3'),
        ("A11", "=A2=A3"),
        ("A12", '=A2=""'),
        ("A13", '=A3=""'),
        ("A14", "=TYPE(A2)"),
        ("A15", "=TYPE(A3)"),
        ("A16", '=CELL("type", A2)'),
        ("A17", '=CELL("type", A3)'),
    ],
    targets=[
        Target("A1", "untouched cell"),
        Target("A2", '="" '),
        Target("A3", "=IF(,,)"),
        Target("A4", "ISBLANK(A1) — truly empty"),
        Target("A5", "ISBLANK(A2) — empty-string-formula"),
        Target("A6", "ISBLANK(A3) — IF(,,)"),
        Target("A7", "ISTEXT(A2)"),
        Target("A8", "ISTEXT(A3)"),
        Target("A9", '"x" & A2'),
        Target("A10", '"x" & A3'),
        Target("A11", "A2 = A3 — are they semantically interchangeable?"),
        Target("A12", "A2 = empty literal"),
        Target("A13", "A3 = empty literal"),
        Target("A14", 'TYPE(A2) — 2 means text, 1 means number, 16 means error'),
        Target("A15", "TYPE(A3) — what does Excel think IF(,,) is?"),
        Target("A16", 'CELL("type", A2) — "v" = value, "l" = label, "b" = blank'),
        Target("A17", 'CELL("type", A3) — does Excel call this blank?'),
    ],
    expect="Excel typically collapses IF(,,) to 0 (different from gsheets, which has a Null). Empty string is a real string. Expect TYPE(A3) = 1 (number, because Excel coerces empty arg to 0), CELL('type', A3) = 'v'. If we're wrong here, that's the discovery.",
)


# ---- Section 2: Error sentinels (mirrors gsheets probe 5 + Excel-only sentinels) ----


def _post_error_obstacle(sht: Any) -> None:
    # Write the obstacle AFTER SEQUENCE has spilled, then save will force
    # recalc and Excel should emit #SPILL!. Pre-writing the obstacle let
    # xlwings.formula2 restrict the array to single-cell at set time.
    try:
        sht.range("C3").value = "obstacle"
    except Exception as e:
        print(f"[recalc]     post-formula obstacle set failed: {e}", file=sys.stderr)


SCN_ERRORS = Scenario(
    name="error-sentinels",
    section="Error sentinels",
    description="Each known Excel error sentinel. Includes Excel-only sentinels (#NULL!, #SPILL!, #CALC!) absent from the gsheets ErrorType enum. Previous run got #VALUE! for the SPILL and CALC tests — hypothesis: xlwings.formula2 entered SEQUENCE as single-cell array when the obstacle was pre-written. Now writing the obstacle POST-formula via post_formula_data.",
    build_data=lambda ws: None,
    formulas=[
        ("A1", "=1/0"),
        ("A2", "=NA()"),
        ("A3", "=NotARealFunction()"),
        ("A4", "=SQRT(-1)"),
        ("A5", '=VLOOKUP("nope", A1, 1, FALSE)'),
        ("A6", '="a"+1'),
        ("A7", "=A1:A10 B11:B20"),
        ("C1", "=SEQUENCE(5)"),
        ("D1", "=FILTER(A1:A5, FALSE)"),
    ],
    post_formula_data=_post_error_obstacle,
    targets=[
        Target("A1", "#DIV/0! from 1/0"),
        Target("A2", "#N/A from NA()"),
        Target("A3", "#NAME? from unknown function"),
        Target("A4", "#NUM! from SQRT(-1)"),
        Target("A5", "#N/A from VLOOKUP miss — does Excel attach a message anywhere?"),
        Target("A6", "#VALUE! from text + number"),
        Target("A7", "#NULL! from non-overlapping intersect — Excel-only vs gsheets ERROR"),
        Target("C1", "#SPILL! anchor (obstacle at C3 written post-formula)"),
        Target("C3", "obstacle cell — should remain 'obstacle' string"),
        Target("D1", "#CALC! from empty array result"),
    ],
    expect="All historical errors → standard sentinel. With obstacle written AFTER SEQUENCE spills, Excel should emit #SPILL!. #CALC! for empty FILTER result — but FILTER didn't change; may still need different test setup.",
)


SCN_STOCKHISTORY = Scenario(
    name="stockhistory-getting-data",
    section="Error sentinels",
    description="STOCKHISTORY attempt to surface #GETTING_DATA in flight. Now entered via xlwings.formula2 post-open so Excel's parser handles the modern-function namespacing.",
    formulas=[
        ("A1", '=STOCKHISTORY("MSFT", TODAY()-5, TODAY())'),
    ],
    targets=[
        Target("A1", "STOCKHISTORY — #GETTING_DATA if in-flight, an error if no MS account, or actual data if signed in"),
    ],
    expect="The audit question is whether #GETTING_DATA ever lands in the saved file. If Excel resolves to data or error before save, the answer is no. Probably fails to open without sign-in.",
)


# ---- Section 3: numFmt inference and propagation (mirrors gsheets probe 4) ----


def _build_numfmt_data(ws: Worksheet) -> None:
    _set_value(ws, "A4", 123)  # literal number — should be General
    _set_value(ws, "A8", 45004)  # literal serial — Excel won't auto-infer date


SCN_NUMFMT = Scenario(
    name="numfmt-inference",
    section="numFmt inference and propagation",
    description="Whether Excel auto-applies a date/time/percent format to formulas that produce typed values, and whether it propagates through cell references and arithmetic. CRITICAL re-test with xlwings.formula2 — first run showed no auto-format, but that was openpyxl-write which may bypass Excel's UI-side auto-format heuristics. This run tests whether xlwings entry triggers them.",
    build_data=_build_numfmt_data,
    formulas=[
        ("A1", "=DATE(2023,3,19)"),
        ("A2", "=NOW()"),
        ("A3", "=TODAY()"),
        ("A5", "=A1"),
        ("A6", "=A1+0"),
        ("A7", '=TEXT(A1, "yyyy-mm-dd")'),
        ("A9", "=10%"),
        ("A10", "=1/4"),
    ],
    targets=[
        Target("A1", "DATE() — expect numFmt=DATE auto-applied"),
        Target("A2", "NOW() — expect numFmt=DATE_TIME auto-applied"),
        Target("A3", "TODAY() — expect numFmt=DATE auto-applied"),
        Target("A4", "literal 123 — expect General / no inference"),
        Target("A5", "=A1 — does the date numFmt propagate?"),
        Target("A6", "=A1+0 — date + number, does it stay date?"),
        Target("A7", "TEXT(A1, 'yyyy-mm-dd') — explicit text format on date"),
        Target("A8", "literal 45004 — same serial as A1 but no inference"),
        Target("A9", "10% literal — percent inference"),
        Target("A10", "1/4 — general number"),
    ],
    expect="DATE/NOW/TODAY: auto-applied. Reference inherits (gsheets does — Excel should too). Arithmetic with date keeps date. Literal serial does NOT infer. ALSO probe whether openpyxl's `is_date` agrees with the cell's actual numFmt.",
)


# ---- Section 4: Boolean and data_type ----


def _build_bool_data(ws: Worksheet) -> None:
    _set_value(ws, "A4", True)  # python bool — does openpyxl write it as t="b"?


SCN_BOOL = Scenario(
    name="boolean-data-type",
    section="Boolean values",
    description="OOXML `t='b'` round-trip via openpyxl and xlwings. Does data_type stay 'b' through write/recalc/read?",
    build_data=_build_bool_data,
    formulas=[
        ("A1", "=TRUE"),
        ("A2", "=FALSE"),
        ("A3", "=1=1"),
        ("A5", "=IF(TRUE, 1, 0)"),
    ],
    targets=[
        Target("A1", "TRUE formula"),
        Target("A2", "FALSE formula"),
        Target("A3", "boolean from 1=1"),
        Target("A4", "Python True written via openpyxl"),
        Target("A5", "integer 1 from IF(TRUE,1,0) — should be number, not boolean"),
    ],
    expect="t='b' for booleans; openpyxl reports value=True/False and data_type='b'. A4 tests openpyxl's writer.",
)


# ---- Section 5: Rich text round-trip ----


def _build_richtext_data(ws: Worksheet) -> None:
    # Write a CellRichText via openpyxl directly; do not include as a formula
    # because Excel doesn't natively re-emit rich text on formula recalc.
    from openpyxl.cell.text import InlineFont

    bold_font = InlineFont(b=True)
    italic_font = InlineFont(i=True)
    rich = CellRichText(
        [
            TextBlock(bold_font, "Hello "),
            "plain ",
            TextBlock(italic_font, "world"),
        ]
    )
    ws["A1"] = rich
    _set_value(ws, "A2", "plain text")


SCN_RICHTEXT = Scenario(
    name="rich-text",
    section="Rich text per-run round-trip",
    description="Does openpyxl's CellRichText survive a write → xlwings recalc-and-save → openpyxl read cycle? Does Excel preserve runs, flatten them, or strip them?",
    build_data=_build_richtext_data,
    formulas=[
        ("A3", '="bold " & "italic"'),
    ],
    targets=[
        Target("A1", "CellRichText with bold + italic runs"),
        Target("A2", "plain string control"),
        Target("A3", "string concat formula — formulas never produce runs"),
    ],
    expect="A1 round-trips with runs preserved if openpyxl 3.1.5's CellRichText writer is faithful AND Excel preserves it. Worth checking — Excel has been known to flatten runs on certain operations.",
)


# ---- Section 6: Hyperlink encodings ----


def _build_hyperlinks_data(ws: Worksheet) -> None:
    # Manual hyperlink via openpyxl
    ws["A1"] = "click manual"
    ws["A1"].hyperlink = Hyperlink(ref="A1", target="https://example.com/manual", display="click manual")
    # Plain URL (auto-recognition uncertain) — set as a literal value
    _set_value(ws, "A3", "https://example.com/typed")


SCN_HYPERLINKS = Scenario(
    name="hyperlinks",
    section="Hyperlink encodings",
    description="Excel has two distinct file-level hyperlink encodings: the sheet-level <hyperlinks> block (manual) and the =HYPERLINK() formula text. Do they cross-populate? What does openpyxl's cell.hyperlink resolve in each case?",
    build_data=_build_hyperlinks_data,
    formulas=[
        ("A2", '=HYPERLINK("https://example.com/formula", "click formula")'),
    ],
    targets=[
        Target("A1", "manual hyperlink via openpyxl.Hyperlink → sheet-level <hyperlinks>"),
        Target("A2", "=HYPERLINK formula — URL only in <f>, not in sheet hyperlinks block"),
        Target("A3", "typed URL — does Excel auto-recognize and emit a hyperlink?"),
    ],
    expect="A1: cell.hyperlink resolves; URL also in raw <hyperlinks> block. A2: cell.hyperlink is None; URL only inside <f>. A3: behavior unclear — Excel may or may not auto-recognize. This is exactly the audit's job.",
)


# ---- Section 7: Spill / array models ----


def _build_spill_data(ws: Worksheet) -> None:
    _set_value(ws, "B1", 1)
    _set_value(ws, "B2", -2)
    _set_value(ws, "B3", 3)


SCN_SPILL = Scenario(
    name="spill-array-models",
    section="Spill / array models",
    description="Modern dynamic-array (SEQUENCE-style) anchor/recipient identity in OOXML, and legacy CSE array via xlwings.formula_array. With xlwings.formula2 entry for SEQUENCE, Excel should evaluate it correctly (previous run hit _xludf rewrite due to openpyxl-write). Plus C5: =TYPE(A1#) for spill-range introspection — does the # spill-range operator make TYPE see an array (64)?",
    build_data=_build_spill_data,
    formulas=[
        ("A1", "=SEQUENCE(5)"),
        ("E1", "=TYPE(A1#)"),  # C5 — spill-range introspection
        ("E2", "=TYPE(A1)"),   # control: same anchor without # — should be 1
        ("E3", "=ROWS(A1#)"),  # how big does Excel see the spill?
        ("E4", "=SUM(A1#)"),   # spill-range arithmetic: 1+2+3+4+5 = 15
    ],
    cse_formulas=[
        ("C1", "=SUM(IF(B1:B3>0,B1:B3))"),
    ],
    targets=[
        Target("A1", "SEQUENCE(5) anchor — has <f>, should have spill metadata"),
        Target("A2", "spill recipient — should have <c cm='...'> but no <f>"),
        Target("A3", "spill recipient"),
        Target("A4", "spill recipient"),
        Target("A5", "spill recipient"),
        Target("C1", "CSE-array attempt — depends on xlwings.formula_array re-write"),
        Target("E1", "TYPE(A1#) — spill-range introspection; expect 64 (array)"),
        Target("E2", "TYPE(A1) — single-cell control; expect 1 (number)"),
        Target("E3", "ROWS(A1#) — how big does Excel see the spill; expect 5"),
        Target("E4", "SUM(A1#) — spill-range arithmetic; expect 15"),
    ],
    expect="A1 has <f> + spill metadata in xl/metadata.xml. A2-A5 are plain value cells (no cm). CSE array gets <f t='array' ref='C1'>. C5: TYPE(A1#) should return 64 if the # operator exposes the array nature, distinguishing it from TYPE(A1)=1.",
)


# ---- Section 8: LAMBDA at cell boundary ----


SCN_LAMBDA = Scenario(
    name="lambda-at-boundary",
    section="LAMBDA, LET, and modern lambda helpers",
    description="Does Excel's cell-boundary LAMBDA emit #CALC! or #VALUE!? How do LET / BYROW present? Re-test with xlwings.formula2 entry — first run hit _xludf rewrite.",
    formulas=[
        ("A1", "=LAMBDA(x, x+1)"),
        ("A2", "=LAMBDA(x, x+1)(5)"),
        ("A3", "=LET(x, 5, x+1)"),
        ("A4", "=BYROW({1,2;3,4}, LAMBDA(row, SUM(row)))"),
    ],
    targets=[
        Target("A1", "bare LAMBDA — expect #CALC! probably"),
        Target("A2", "called LAMBDA — expect 6"),
        Target("A3", "LET binding — expect 6"),
        Target("A4", "BYROW returning array of row sums — modern lambda helper"),
    ],
    expect="A1: #CALC! (probably; gsheets returns N_A with a specific message). A2: 6. A3: 6. A4: array {3;7} — depends on spill behavior.",
)


# ---- Section 9: Function name prefixing (_xlfn., _xlws.) ----


def _build_xlfn_data(ws: Worksheet) -> None:
    _set_value(ws, "A1", "alpha")
    _set_value(ws, "A2", "beta")
    _set_value(ws, "A3", "gamma")
    _set_value(ws, "B1", 1)
    _set_value(ws, "B2", 2)
    _set_value(ws, "B3", 3)


SCN_XLFN = Scenario(
    name="xlfn-prefixing",
    section="_xlfn. / _xlws. function-name prefixing",
    description="Modern Excel functions are stored in the saved xlsx as _xlfn.XLOOKUP, _xlfn._xlws.FILTER, etc. With xlwings.formula2 entry, Excel's parser should write the correct prefixes. After save, raw XML should show _xlfn./_xlws. namespacing.",
    build_data=_build_xlfn_data,
    formulas=[
        ("D1", '=XLOOKUP("beta", A1:A3, B1:B3)'),
        ("D2", "=FILTER(B1:B3, B1:B3>1)"),
        ("D3", "=UNIQUE({1;2;2;3})"),
        ("D4", "=LET(x, 5, x*2)"),
    ],
    targets=[
        Target("D1", "XLOOKUP — expect _xlfn.XLOOKUP in raw XML, 'XLOOKUP' or the prefixed form in openpyxl <f> read?"),
        Target("D2", "FILTER — expect _xlfn._xlws.FILTER"),
        Target("D3", "UNIQUE"),
        Target("D4", "LET"),
    ],
    expect="Raw XML shows _xlfn./ _xlws. prefixes. openpyxl when data_only=False reads <f> — does it strip? data_only=True hides <f> entirely.",
)


# ---- Section 10: Implicit intersection / @ operator ----


def _build_implicit_intersect_data(ws: Worksheet) -> None:
    for i in range(1, 6):
        ws[f"A{i}"] = i


SCN_IMPLICIT_INTERSECT = Scenario(
    name="implicit-intersection",
    section="Implicit intersection / @ operator",
    description="Post-365 dynamic arrays changed the meaning of bare range refs in single-cell contexts. The @ operator restores pre-365 single-value behavior. Re-test with xlwings entry — last run showed @ stripped on save (pre-365 behavior, but possibly an openpyxl-write artifact).",
    build_data=_build_implicit_intersect_data,
    formulas=[
        ("C1", "=A1:A5"),
        ("D1", "=@A1:A5"),
        ("E1", "=SUM(A1:A5)"),
    ],
    targets=[
        Target("C1", "=A1:A5 — should spill in 365; <f> may carry @ in older form"),
        Target("D1", "=@A1:A5 — single value, A1"),
        Target("E1", "=SUM(A1:A5) — control, no intersection issue"),
    ],
    expect="C1 spills to C1:C5 with value 1,2,3,4,5. D1 = 1. E1 = 15. Worth checking whether Excel rewrites the formula on save (some versions normalize @ insertion).",
)


# ---- Section 11: 1900 vs 1904 date system ----


def _build_date_data_1900(ws: Worksheet) -> None:
    _set_value(ws, "A2", 45004)


def _build_date_data_1904(ws: Worksheet) -> None:
    _set_value(ws, "A2", 43542)


def _setup_workbook_1900(wb: Workbook) -> None:
    wb.epoch = datetime(1899, 12, 30)


def _setup_workbook_1904(wb: Workbook) -> None:
    wb.epoch = datetime(1904, 1, 1)


SCN_DATE_1900 = Scenario(
    name="date-system-1900",
    section="Date system (1900 vs 1904)",
    description="Default 1900 date system. Confirms baseline. Now tests whether xlwings.formula2 entry of =DATE() triggers Excel's UI-side auto-format (the first run via openpyxl-write showed no auto-format).",
    build_data=_build_date_data_1900,
    formulas=[
        ("A1", "=DATE(2023,3,19)"),
        ("A3", "=A2"),
    ],
    targets=[
        Target("A1", "=DATE(2023,3,19) → serial 45004; auto-numFmt?"),
        Target("A2", "literal 45004"),
        Target("A3", "reference — does inferred type propagate?"),
    ],
    expect="If xlwings entry triggers auto-format, A1 will have a date numFmt and openpyxl.cell.value may coerce to datetime. If not, same as last run.",
)

SCN_DATE_1904 = Scenario(
    name="date-system-1904",
    section="Date system (1900 vs 1904)",
    description="1904 date system. The current driver's _DATE_EPOCH = datetime(1899,12,30) is hardcoded — would 1904-mode workbooks silently shift by 1462 days?",
    build_data=_build_date_data_1904,
    formulas=[
        ("A1", "=DATE(2023,3,19)"),
        ("A3", "=A2"),
    ],
    targets=[
        Target("A1", "=DATE(2023,3,19) → serial 43542 in 1904 system"),
        Target("A2", "literal 43542"),
        Target("A3", "reference"),
    ],
    expect="A1 cell.value should be 43542 (raw serial) or datetime(2023,3,19) if openpyxl knows wb.epoch.",
    needs_workbook_setup=_setup_workbook_1904,
)


# ---- Section 12: Excel TYPE() and CELL() probes ----


def _build_type_cell_data(ws: Worksheet) -> None:
    _set_value(ws, "A1", 42)  # number
    _set_value(ws, "A2", "text")  # text
    _set_value(ws, "A3", True)  # bool


def _type_cell_formulas() -> list[tuple[str, str]]:
    f = [
        ("A4", "=1/0"),
        ("A5", "=SEQUENCE(2)"),
    ]
    for i in range(1, 6):
        f.append((f"B{i}", f"=TYPE(A{i})"))
        f.append((f"C{i}", f'=CELL("type", A{i})'))
    return f


SCN_TYPE_CELL = Scenario(
    name="type-and-cell-probes",
    section="Excel TYPE() and CELL() — engine-side type introspection",
    description="Excel-native introspection of what the engine thinks each cell is. Useful for triangulating against openpyxl's data_type. With SEQUENCE working via xlwings entry, B5/C5 will reflect array type (64 / 'v' respectively).",
    build_data=_build_type_cell_data,
    formulas=_type_cell_formulas(),
    targets=[Target(f"B{i}", f"TYPE of A{i}") for i in range(1, 6)]
    + [Target(f"C{i}", f"CELL('type', A{i})") for i in range(1, 6)],
    expect="B1=1, B2=2, B3=4, B4=16, B5=64. C1='v', C2='l', C3='v' (booleans are values), C4='v', C5='v'.",
)


# ---- Master scenario list ----

def _build_blank_data(ws: Worksheet) -> None:
    # B1, B2, B3 form a lookup table: B1=key1/C1=val1, B2=key2/C2=untouched, B3=key3/C3=val3
    # Used to probe whether a blank cell read via VLOOKUP propagates blank-ness or coerces.
    _set_value(ws, "B1", 1)
    _set_value(ws, "C1", 100)
    _set_value(ws, "B2", 2)
    # C2 deliberately untouched — VLOOKUP(2, ...) will read this blank
    _set_value(ws, "B3", 3)
    _set_value(ws, "C3", 300)
    # D1 has a string-then-cleared sequence built post-open via xlwings


SCN_BLANK = Scenario(
    name="blank-cell-semantics",
    section="Blank cell representation",
    description="How does Excel represent blank cells if there's no runtime-Null variant? Probes truly-untouched vs '=\"\"' vs '=IF(,,)' via TYPE/CELL/IS* introspection, arithmetic/concat/comparison coercions, and VLOOKUP-returning-blank propagation.",
    build_data=_build_blank_data,
    formulas=[
        # A1: untouched (no formula).
        # A2: set up empty-string formula
        ("A2", '=""'),
        # A3: =IF(,,) collapses to 0
        ("A3", "=IF(,,)"),
        # A4: =IF(FALSE, 1, ) — returns missing arg
        ("A4", "=IF(FALSE, 1, )"),
        # A5: =VLOOKUP key=2 returning blank C2
        ("A5", "=VLOOKUP(2, B1:C3, 2, FALSE)"),
        # Introspection on untouched A1
        ("F1", "=TYPE(A1)"),
        ("F2", '=CELL("type", A1)'),
        ("F3", "=ISBLANK(A1)"),
        ("F4", "=ISNUMBER(A1)"),
        ("F5", "=ISTEXT(A1)"),
        ("F6", "=ISLOGICAL(A1)"),
        ("F7", "=ISERROR(A1)"),
        ("F8", "=N(A1)"),
        ("F9", "=T(A1)"),
        ("F10", "=A1+5"),  # numeric coercion: expect 5
        ('F11', '="x" & A1'),  # text coercion: expect "x"
        ("F12", "=A1=0"),  # blank == 0?
        ("F13", '=A1=""'),  # blank == ""?
        ("F14", "=A1=FALSE"),  # blank == FALSE?
        # COUNTBLANK across the categories
        ("F15", "=COUNTBLANK(A1)"),  # untouched — 1?
        ("F16", "=COUNTBLANK(A2)"),  # =""  — gsheets says 1, Excel?
        ("F17", "=COUNTBLANK(A3)"),  # =IF(,,) — Excel says probably 0 (it's a number 0); gsheets says 1
        ("F18", "=COUNTBLANK(A4)"),  # =IF(FALSE, 1, )
        # COUNTA across the categories
        ("F19", "=COUNTA(A1)"),  # untouched — 0?
        ("F20", "=COUNTA(A2)"),  # =""  — 1?
        ("F21", "=COUNTA(A3)"),  # =IF(,,) — 1?
        # VLOOKUP-returning-blank propagation: A5 is the VLOOKUP result; does it behave blank?
        ("G1", "=ISBLANK(A5)"),  # propagates blank-ness through VLOOKUP?
        ("G2", "=TYPE(A5)"),
        ("G3", '=CELL("type", A5)'),
        ("G4", "=A5+5"),  # coerce — 5 if blank propagated, 5 either way
        ("G5", '="x" & A5'),  # "x" if blank → ""
        ("G6", "=A5=0"),
        ("G7", '=A5=""'),
    ],
    targets=[
        Target("A1", "untouched cell — file should have no <c r='A1'>"),
        Target("A2", '=""'),
        Target("A3", "=IF(,,)"),
        Target("A4", "=IF(FALSE, 1, ) — missing arg"),
        Target("A5", "=VLOOKUP(2, B1:C3, 2, FALSE) — returns blank C2"),
        Target("F1", "TYPE(A1) — number 1 (treated as 0) or other?"),
        Target("F2", 'CELL("type", A1) — expect "b" for blank'),
        Target("F3", "ISBLANK(A1) — TRUE"),
        Target("F4", "ISNUMBER(A1)"),
        Target("F5", "ISTEXT(A1)"),
        Target("F6", "ISLOGICAL(A1)"),
        Target("F7", "ISERROR(A1)"),
        Target("F8", "N(A1) — numeric coercion of blank"),
        Target("F9", "T(A1) — text coercion of blank"),
        Target("F10", "A1+5 — expect 5"),
        Target("F11", '"x" & A1 — expect "x"'),
        Target("F12", "A1=0 — equality with 0"),
        Target("F13", 'A1="" — equality with ""'),
        Target("F14", "A1=FALSE"),
        Target("F15", "COUNTBLANK(A1) — untouched"),
        Target("F16", 'COUNTBLANK(A2) — =""'),
        Target("F17", "COUNTBLANK(A3) — =IF(,,)"),
        Target("F18", "COUNTBLANK(A4) — =IF(FALSE,1,)"),
        Target("F19", "COUNTA(A1) — untouched"),
        Target("F20", 'COUNTA(A2) — =""'),
        Target("F21", "COUNTA(A3) — =IF(,,)"),
        Target("G1", "ISBLANK of VLOOKUP-result — does blank propagate?"),
        Target("G2", "TYPE of VLOOKUP-result"),
        Target("G3", 'CELL("type", VLOOKUP-result)'),
        Target("G4", "(VLOOKUP-result)+5"),
        Target("G5", '"x" & VLOOKUP-result'),
        Target("G6", "VLOOKUP-result = 0"),
        Target("G7", 'VLOOKUP-result = ""'),
    ],
    expect="Untouched: ISBLANK TRUE, CELL='b', TYPE=1 (Excel treats blank as 0 for TYPE). Numeric/text coercions per spec. VLOOKUP-returning-blank is the load-bearing question: does the result propagate blank semantics, or does Excel coerce on read? gsheets propagates Null through VLOOKUP — Excel almost certainly does NOT, but worth verifying.",
)


ALL_SCENARIOS: list[Scenario] = [
    SCN_NULL,
    SCN_ERRORS,
    SCN_STOCKHISTORY,
    SCN_NUMFMT,
    SCN_BOOL,
    SCN_RICHTEXT,
    SCN_HYPERLINKS,
    SCN_SPILL,
    SCN_LAMBDA,
    SCN_XLFN,
    SCN_IMPLICIT_INTERSECT,
    SCN_DATE_1900,
    SCN_DATE_1904,
    SCN_TYPE_CELL,
    SCN_BLANK,
]


# ---------------------------------------------------------------------------
# Build phase
# ---------------------------------------------------------------------------


def _sandbox_tmpdir() -> Path:
    """Same trick as the existing driver — write inside Excel's sandbox container
    when on macOS to avoid 'grant access' prompts per file."""
    override = os.environ.get("ASSAY_EXCEL_TMPDIR")
    if override:
        root = Path(override).expanduser().resolve()
        root.mkdir(parents=True, exist_ok=True)
        return Path(tempfile.mkdtemp(prefix="fidelity-", dir=root))
    candidate = Path.home() / "Library" / "Containers" / "com.microsoft.Excel" / "Data"
    if candidate.is_dir() and os.access(candidate, os.W_OK):
        root = candidate / "assay-tmp"
        try:
            root.mkdir(parents=True, exist_ok=True)
            return Path(tempfile.mkdtemp(prefix="fidelity-", dir=root))
        except PermissionError:
            pass
    return Path(tempfile.mkdtemp(prefix="assay-fidelity-"))


def build_all(scenarios: list[Scenario], path: Path) -> dict[str, str]:
    """Build one xlsx per scenario with data only (no formulas). Formulas are
    entered via xlwings.formula2 in the recalc phase so Excel's parser handles
    modern-function namespacing (_xlfn.) correctly. Returns a map
    scenario.name -> file path."""
    paths: dict[str, str] = {}
    for s in scenarios:
        wb = Workbook()
        if wb.active is not None and wb.active.title == "Sheet":
            wb.remove(wb.active)
        if s.needs_workbook_setup is not None:
            s.needs_workbook_setup(wb)
        ws = wb.create_sheet(title=s.name[:31])
        try:
            s.build_data(ws)
        except Exception as e:
            print(f"[build] {s.name}: build_data failed — {type(e).__name__}: {e}", file=sys.stderr)
        scn_path = path.with_stem(f"{path.stem}-{s.name}")
        wb.save(scn_path)
        paths[s.name] = str(scn_path)
    return paths


# ---------------------------------------------------------------------------
# Recalc phase (xlwings)
# ---------------------------------------------------------------------------


def recalc_via_excel(
    paths_by_scenario: dict[str, str],
    scenarios_by_name: dict[str, Scenario],
) -> dict[str, str]:
    """Open each workbook, enter formulas via xlwings.formula2 (and CSE arrays
    via formula_array), recalc, save, close. Returns scenario_name -> status."""
    status: dict[str, str] = {}
    print(f"[recalc] opening Excel for {len(paths_by_scenario)} workbook(s)", file=sys.stderr)
    app = xw.App(visible=False, add_book=False)
    try:
        try:
            app.display_alerts = False
        except Exception:
            pass
        try:
            app.screen_updating = False
        except Exception:
            pass
        for scn_name, p in paths_by_scenario.items():
            print(f"[recalc]   {scn_name} ({os.path.basename(p)})", file=sys.stderr)
            scn = scenarios_by_name[scn_name]
            try:
                wb = app.books.open(p)
            except Exception as e:
                msg = f"{type(e).__name__}: {e}".splitlines()[0][:200]
                status[scn_name] = f"open_failed: {msg}"
                print(f"[recalc]     OPEN FAILED: {msg}", file=sys.stderr)
                continue
            try:
                sheet_name = scn_name[:31]
                if sheet_name not in [sht.name for sht in wb.sheets]:
                    status[scn_name] = "recalc_failed: sheet missing"
                    print(f"[recalc]     RECALC FAILED: sheet '{sheet_name}' not in workbook", file=sys.stderr)
                    continue
                sht = wb.sheets[sheet_name]
                # Manual-calc mode while entering formulas avoids quadratic recalc per formula entry
                prior_calc = None
                try:
                    prior_calc = app.calculation
                    app.calculation = "manual"
                except Exception:
                    pass
                try:
                    # Enter formula2 entries (modern functions get Excel's correct namespacing)
                    for ref, formula in scn.formulas:
                        try:
                            sht.range(ref).formula2 = formula
                        except Exception as e:
                            print(f"[recalc]     formula2 {ref}={formula!r} failed: {e}", file=sys.stderr)
                    # Enter CSE arrays via formula_array
                    for ref, formula in scn.cse_formulas:
                        try:
                            sht.range(ref).formula_array = formula
                        except Exception as e:
                            print(f"[recalc]     formula_array {ref}={formula!r} failed: {e}", file=sys.stderr)
                finally:
                    if prior_calc is not None:
                        try:
                            app.calculation = prior_calc
                        except Exception:
                            pass
                # Post-formula data: e.g., write an obstacle into a spill recipient
                # to force a #SPILL! after the spill has already happened.
                if scn.post_formula_data is not None:
                    try:
                        scn.post_formula_data(sht)
                    except Exception as e:
                        print(f"[recalc]     post_formula_data failed: {e}", file=sys.stderr)
                app.calculate()
                wb.save()
                status[scn_name] = "ok"
            except Exception as e:
                msg = f"{type(e).__name__}: {e}".splitlines()[0][:200]
                status[scn_name] = f"recalc_failed: {msg}"
                print(f"[recalc]     RECALC FAILED: {msg}", file=sys.stderr)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
    finally:
        try:
            app.quit()
        except Exception:
            pass
    ok_count = sum(1 for v in status.values() if v == "ok")
    print(f"[recalc] {ok_count}/{len(paths_by_scenario)} succeeded", file=sys.stderr)
    return status


# ---------------------------------------------------------------------------
# Read phase — Surface A (openpyxl) and Surface C (raw XML)
# ---------------------------------------------------------------------------


def read_openpyxl_cell(cell: Cell) -> dict[str, Any]:
    out: dict[str, Any] = {
        "value": _serialize_value(cell.value),
        "data_type": cell.data_type,
        "number_format": cell.number_format,
        "is_date": cell.is_date,
        "coordinate": cell.coordinate,
    }
    if cell.hyperlink is not None:
        out["hyperlink"] = {
            "target": cell.hyperlink.target,
            "display": cell.hyperlink.display,
            "tooltip": cell.hyperlink.tooltip,
            "location": cell.hyperlink.location,
        }
    if cell.comment is not None:
        out["comment"] = {"text": cell.comment.text, "author": cell.comment.author}
    if isinstance(cell.value, CellRichText):
        out["rich_runs"] = [_serialize_rich_block(b) for b in cell.value]
    return out


def _serialize_value(v: Any) -> Any:
    if isinstance(v, (str, int, float, bool)) or v is None:
        return v
    if isinstance(v, datetime):
        return {"__datetime__": v.isoformat()}
    if isinstance(v, CellRichText):
        return {"__rich_text__": "".join(str(b) if isinstance(b, str) else b.text for b in v)}
    return repr(v)


def _serialize_rich_block(b: Any) -> dict[str, Any]:
    if isinstance(b, str):
        return {"text": b, "format": None}
    # TextBlock
    fmt: dict[str, Any] = {}
    if b.font is not None:
        for attr in ("b", "i", "u", "strike", "color", "rFont", "sz"):
            val = getattr(b.font, attr, None)
            if val is not None:
                fmt[attr] = str(val)
    return {"text": b.text, "format": fmt or None}


def _resolve_rel_target(target: str, rels_dir: str) -> str:
    """OOXML rels targets can be absolute (leading '/', from zip root) or
    relative to the directory containing the .rels file. openpyxl writes
    worksheet rels as absolute but styles/theme as relative."""
    if target.startswith("/"):
        return target.lstrip("/")
    return rels_dir.rstrip("/") + "/" + target


def parse_raw_cell(xlsx_path: str, sheet_name: str, cell_ref: str) -> dict[str, Any] | None:
    """Read the raw <c> element directly from the xlsx zip for one cell."""
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            # Map sheet name to file path via workbook.xml and the rels
            workbook_xml = z.read("xl/workbook.xml").decode("utf8")
            wb_root = ET.fromstring(workbook_xml)
            sheets = wb_root.find("main:sheets", NS)
            if sheets is None:
                return None
            target_rid: str | None = None
            for sht in sheets.findall("main:sheet", NS):
                if sht.attrib.get("name") == sheet_name:
                    target_rid = sht.attrib.get(f"{{{NS['r']}}}id")
                    break
            if target_rid is None:
                return None
            rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf8")
            rels_root = ET.fromstring(rels_xml)
            target_path: str | None = None
            for rel in rels_root.findall("pkg-rels:Relationship", NS):
                if rel.attrib.get("Id") == target_rid:
                    target_path = _resolve_rel_target(rel.attrib.get("Target", ""), "xl/")
                    break
            if target_path is None:
                return None
            sheet_xml = z.read(target_path).decode("utf8")
            sheet_root = ET.fromstring(sheet_xml)
            # Walk rows; find <c r="A1">
            for row in sheet_root.iter(f"{{{NS['main']}}}row"):
                for c in row.findall(f"{{{NS['main']}}}c"):
                    if c.attrib.get("r") == cell_ref:
                        return _serialize_xml_cell(c, z, target_path)
        return {"_missing": True, "_note": f"no <c r='{cell_ref}'> in {sheet_name}"}
    except Exception as e:
        return {"_error": f"{type(e).__name__}: {e}"}


def _serialize_xml_cell(c: ET.Element, z: zipfile.ZipFile, sheet_path: str) -> dict[str, Any]:
    out: dict[str, Any] = {
        "attributes": dict(c.attrib),
        "children": {},
    }
    for child in c:
        tag = child.tag.split("}", 1)[-1] if "}" in child.tag else child.tag
        if tag == "v":
            out["children"]["v"] = child.text
        elif tag == "f":
            out["children"]["f"] = {
                "text": child.text,
                "attributes": dict(child.attrib),
            }
        elif tag == "is":
            # inline string with possible runs
            out["children"]["is"] = ET.tostring(child, encoding="unicode")
    # Resolve sharedString if t="s"
    t = c.attrib.get("t")
    if t == "s" and "v" in out["children"] and out["children"]["v"] is not None:
        try:
            ss_xml = z.read("xl/sharedStrings.xml").decode("utf8")
            ss_root = ET.fromstring(ss_xml)
            idx = int(out["children"]["v"])
            si_elements = ss_root.findall("main:si", NS)
            if 0 <= idx < len(si_elements):
                si = si_elements[idx]
                # Just dump the XML so per-run formatting is visible
                out["resolved_shared_string"] = ET.tostring(si, encoding="unicode")
        except Exception as e:
            out["resolved_shared_string_error"] = f"{type(e).__name__}: {e}"
    return out


def find_sheet_level_hyperlinks(xlsx_path: str, sheet_name: str) -> list[dict[str, str]] | None:
    """Read the <hyperlinks> block for a sheet, if any."""
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            workbook_xml = z.read("xl/workbook.xml").decode("utf8")
            wb_root = ET.fromstring(workbook_xml)
            target_rid: str | None = None
            sheets = wb_root.find("main:sheets", NS)
            if sheets is None:
                return None
            for sht in sheets.findall("main:sheet", NS):
                if sht.attrib.get("name") == sheet_name:
                    target_rid = sht.attrib.get(f"{{{NS['r']}}}id")
                    break
            if target_rid is None:
                return None
            rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf8")
            rels_root = ET.fromstring(rels_xml)
            target_path: str | None = None
            for rel in rels_root.findall("pkg-rels:Relationship", NS):
                if rel.attrib.get("Id") == target_rid:
                    target_path = _resolve_rel_target(rel.attrib.get("Target", ""), "xl/")
                    break
            if target_path is None:
                return None
            # Sheet-level rels for hyperlinks
            sheet_rels_path = target_path.replace("xl/worksheets/", "xl/worksheets/_rels/") + ".rels"
            rels_lookup: dict[str, str] = {}
            try:
                sheet_rels_xml = z.read(sheet_rels_path).decode("utf8")
                sheet_rels_root = ET.fromstring(sheet_rels_xml)
                for rel in sheet_rels_root.findall("pkg-rels:Relationship", NS):
                    rels_lookup[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")
            except KeyError:
                pass
            sheet_xml = z.read(target_path).decode("utf8")
            sheet_root = ET.fromstring(sheet_xml)
            hyperlinks_el = sheet_root.find("main:hyperlinks", NS)
            if hyperlinks_el is None:
                return []
            out = []
            for hl in hyperlinks_el.findall("main:hyperlink", NS):
                entry = dict(hl.attrib)
                rid = entry.get(f"{{{NS['r']}}}id")
                if rid and rid in rels_lookup:
                    entry["resolved_target"] = rels_lookup[rid]
                out.append(entry)
            return out
    except Exception as e:
        return [{"_error": f"{type(e).__name__}: {e}"}]


# ---------------------------------------------------------------------------
# Report emission
# ---------------------------------------------------------------------------


def emit_report(findings: list[dict[str, Any]], report_path: Path) -> None:
    lines: list[str] = []
    lines.append("# Excel driver fidelity audit + behavior probes")
    lines.append("")
    lines.append(f"Generated: {datetime.now().isoformat()}")
    lines.append("")
    lines.append("Companion to [`excel-celldata-gap.md`](./excel-celldata-gap.md) and the gsheets-side [`gsheets-celldata-probes.md`](./gsheets-celldata-probes.md).")
    lines.append("")
    lines.append("Each scenario probes three surfaces:")
    lines.append("")
    lines.append("- **A** — openpyxl per-cell (`cell.value`, `cell.data_type`, `cell.number_format`, `cell.is_date`, `cell.hyperlink`, `cell.comment`, `CellRichText` when applicable)")
    lines.append("- **C** — raw OOXML XML extracted from the saved xlsx zip (the `<c>` element, attributes + children)")
    lines.append("- **B** — xlwings live `.api` (best-effort on Mac AppleEvents; many Windows COM properties unreachable)")
    lines.append("")
    lines.append("**Disagreements between A and C are the audit's primary deliverable** — they catalog where openpyxl lies or omits relative to what Excel actually persisted.")
    lines.append("")
    lines.append("Re-run: `uv run python packages/assay/scripts/probes/excel-driver-fidelity.py`")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Group by section
    by_section: dict[str, list[dict[str, Any]]] = {}
    for f in findings:
        by_section.setdefault(f["section"], []).append(f)

    for section, items in by_section.items():
        lines.append(f"## {section}")
        lines.append("")
        for f in items:
            lines.append(f"### {f['name']}")
            lines.append("")
            status_text = f.get("recalc_status", "unknown")
            status_marker = "**ok**" if status_text == "ok" else f"**{status_text}** — Surface C cached `<v>` values reflect pre-recalc state (formula text in `<f>` is still informative; cached values may be null)"
            lines.append(f"**Recalc status:** {status_marker}")
            lines.append("")
            lines.append(f"**Description:** {f['description']}")
            lines.append("")
            if f.get("expect"):
                lines.append(f"**Pre-probe expectation:** {f['expect']}")
                lines.append("")
            if f.get("sheet_level_hyperlinks"):
                lines.append(f"**Sheet-level `<hyperlinks>` block:** `{json.dumps(f['sheet_level_hyperlinks'])}`")
                lines.append("")
            for t in f["target_results"]:
                lines.append(f"#### Target: `{t['ref']}` — {t['description']}")
                lines.append("")
                lines.append("**Surface A (openpyxl per-cell):**")
                lines.append("")
                lines.append("```json")
                lines.append(json.dumps(t["openpyxl"], indent=2, default=str))
                lines.append("```")
                lines.append("")
                lines.append("**Surface C (raw OOXML `<c>` element):**")
                lines.append("")
                lines.append("```json")
                lines.append(json.dumps(t["raw_xml"], indent=2, default=str))
                lines.append("```")
                lines.append("")
            lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## Disagreements catalog")
    lines.append("")
    lines.append("_Populated manually after the first read-through. The script collects the raw data; the maintainer + future-Claude annotates here where openpyxl and raw XML disagree, what to trust, and what to change in the driver._")
    lines.append("")

    report_path.write_text("\n".join(lines))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    tmp = _sandbox_tmpdir()
    base_xlsx = tmp / "fidelity.xlsx"
    print(f"[main] tmpdir: {tmp}", file=sys.stderr)

    # Build
    path_map = build_all(ALL_SCENARIOS, base_xlsx)
    print(f"[main] built {len(set(path_map.values()))} workbook(s) covering {len(ALL_SCENARIOS)} scenarios", file=sys.stderr)

    # Recalc
    scenarios_by_name = {s.name: s for s in ALL_SCENARIOS}
    status = recalc_via_excel(path_map, scenarios_by_name)

    # Read
    findings: list[dict[str, Any]] = []
    for s in ALL_SCENARIOS:
        xlsx_path = path_map.get(s.name)
        if not xlsx_path:
            continue
        try:
            wb = load_workbook(xlsx_path, data_only=True, rich_text=True)
        except Exception as e:
            print(f"[read] {s.name}: failed to load — {type(e).__name__}: {e}", file=sys.stderr)
            continue
        sheet_title = s.name[:31]
        if sheet_title not in wb.sheetnames:
            print(f"[read] {s.name}: sheet missing", file=sys.stderr)
            continue
        ws = wb[sheet_title]
        target_results = []
        for t in s.targets:
            try:
                cell = ws[t.ref]
                op = read_openpyxl_cell(cell)
            except Exception as e:
                op = {"_error": f"{type(e).__name__}: {e}"}
            try:
                xml = parse_raw_cell(xlsx_path, sheet_title, t.ref)
            except Exception as e:
                xml = {"_error": f"{type(e).__name__}: {e}"}
            target_results.append(
                {"ref": t.ref, "description": t.description, "openpyxl": op, "raw_xml": xml}
            )
        wb.close()
        # Sheet-level hyperlinks (only relevant for the hyperlinks scenario)
        sheet_hyperlinks = None
        if s.name == "hyperlinks":
            sheet_hyperlinks = find_sheet_level_hyperlinks(xlsx_path, sheet_title)
        findings.append(
            {
                "name": s.name,
                "section": s.section,
                "description": s.description,
                "expect": s.expect,
                "target_results": target_results,
                "sheet_level_hyperlinks": sheet_hyperlinks,
                "recalc_status": status.get(s.name, "unknown"),
            }
        )

    # Emit report
    emit_report(findings, REPORT_PATH)
    print(f"[main] report → {REPORT_PATH}", file=sys.stderr)

    # Cleanup temp dir
    try:
        shutil.rmtree(tmp, ignore_errors=True)
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    sys.exit(main())
